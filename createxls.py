import openpyxl
import os

cur_dir = os.getcwd()
file_name = 'M010-MSI-INM-13-ICD - Doha Metro SCADA IO LIST SCD-CIVS (stations_in_1).xlsx'
file_new = 'image_autogen.xlsx'
file_sour = cur_dir + '\\' + file_name
file_new_sour = cur_dir + '\\' + file_new
dict_sys = {'eqpt_list_VHTS': 0, 'eqpt_list_PHE': 1, 'eqpt_list_BACS': 2, 'eqpt_list_CHILLER': 3, 'eqpt_list_CHWH': 4, 'eqpt_list_HVAC': 5}
dict_eq = {'FTC_PUMP': 'DohBacsSwPump_layer', 'SW_PUMP': 'DohBacsSwPump_layer', 'CW_PUMP': 'DohBacsCwPump_layer', 'IR_PUMP': 'DohBacsIrPump_layer', 'TRWS': 'DohBacsTrws_layer', 'WM': 'DohBacsWm_layer', 'LS': 'DohBacsLs_layer', 'LIFT': 'DohBacsLift_layer', 'TRAV_ESC_1': 'DohBacsTravEscEscalator_layer', 'TRAV_ESC_2': 'DohBacsTravEscTravellator_layer', 'DIAG_RIO': 'DohBacsDiagRio_layer', 'DIAG_NWSW': 'DohBacsDiagNwsw_layer','DIAG_PLC': 'DohBacsDiagPlc_layer','DIAG_SER': 'DohBacsDiagSer_layer', 'CPM': 'DohBacsCpm_layer','CHWP': 'DohBacsChwp_layer', 'PMU': 'DohBacsPmu_layer', 'CDS': 'DohBacsCds_layer', 'CHWH': 'DohBacsChwh_layer', 'HEX': 'DohBacsHex_layer', 'PEU': 'DohBacsPeu_layer', 'VRV': 'DohBacsVrv_layer', 'FCU': 'DohBacsFcu_layer', 'CRAC': 'DohBacsCrac_layer', 'ACS_1': 'DohBacsAcsSingle_layer', 'ACS_2': 'DohBacsAcsDual_layer', 'ACMODAMP': 'DohBacsAcModamp_layer'}
dict_id = {'FTC_PUMP': 'BasicEquipment_DohBacsFtcPump_EVqHEL3nEeeHE768d9D71g_1509418578433',  'SW_PUMP': 'BasicEquipment_DohBacsSwPump_8fpicL1JEeekaonQy-nY5g_1509351094807', 'CW_PUMP': 'BasicEquipment_DohBacsCwPump_3jnqYLo1Eee4FZNjhxyf2w_1509012618246', 'IR_PUMP': 'BasicEquipment_DohBacsIrPump_HNYowLrfEeeUOKN4pirLLA_1509085308236', 'TRWS': 'BasicEquipment_DohBacsTrws_rn29YLrBEee2yfXdVkK8vg_1509072667702', 'WM': 'BasicEquipment_DohBacsWm_ZhZJcL1BEeeMG7pPixeMOg_1509347424135', 'LS': 'BasicEquipment_DohBacsLs_9GgFcMT8Eee2FLSNuBaW1Q_1510197636935', 'LIFT': 'BasicEquipment_DohBacsLift_3yJJ4MLNEeetjoi2WK0QMQ_1509957512574', 'TRAV_ESC_1': 'BasicEquipment_DohBacsTravEsc_FZrJ0MK-Eee62KrShjok0Q_1509950732013', 'TRAV_ESC_2': 'BasicEquipment_DohBacsTravEsc_FZrJ0MK-Eee62KrShjok0Q_1509950732013', 'DIAG_RIO': 'BasicEquipment_DohBacsDiagRio_BSA2gLrfEeeUOKN4pirLLA_1509085268456', 'DIAG_NWSW': 'BasicEquipment_DohBacsDiagNwsw_1bxNELo1Eee4FZNjhxyf2w_1509012604001','DIAG_PLC': 'BasicEquipment_DohBacsDiagPlc_66ly8L8nEee5TNQefPaSHQ_1509556383647','DIAG_SER': 'BasicEquipment_DohBacsDiagSer_78GJcL1JEeekaonQy-nY5g_1509351091079', 'CPM': 'BasicEquipment_DohBacsCpm_vyjcMLrBEee2yfXdVkK8vg_1509072695667', 'CHWP': 'BasicEquipment_DohBacsChwp_XUuHMMLLEeeL65l03tuNYA_1509956435748', 'PMU': 'BasicEquipment_DohBacsPmu_Hx6MgMK-Eee62KrShjok0Q_1509950747976', 'CDS': 'BasicEquipment_DohBacsCds_5O-7UMLNEeetjoi2WK0QMQ_1509957522309', 'CHWH': 'BasicEquipment_DohBacsChwh_Ysi1wMLLEeeL65l03tuNYA_1509956444956', 'HEX': 'BasicEquipment_DohBacsHex_yb5dALrBEee2yfXdVkK8vg_1509072713424', 'PEU': 'BasicEquipment_DohBacsPeu_tv7dILrBEee2yfXdVkK8vg_1509072681971', 'VRV': 'BasicEquipment_DohBacsVrv_FTr8MLrfEeeUOKN4pirLLA_1509085295475', 'FCU': 'BasicEquipment_DohBacsFcu_GOmsULrfEeeUOKN4pirLLA_1509085301653', 'CRAC': 'BasicEquipment_DohBacsCrac_ETayALrfEeeUOKN4pirLLA_1509085288736', 'ACS_1': 'BasicEquipment_DohBacsAcs_FY_MUL3nEeeHE768d9D71g_1509418585493', 'ACS_2': 'BasicEquipment_DohBacsAcs_FY_MUL3nEeeHE768d9D71g_1509418585493', 'ACMODAMP': 'BasicEquipment_DohBacsAcModamp_6U7UYLo1Eee4FZNjhxyf2w_1509012636838'}
dict_rd = {'FTC_PUMP': '_yAuL0L_CEeeok6BbH_neNQ',  'SW_PUMP': '_jy7a4L4WEeeBbKd0rZ4ZWQ', 'CW_PUMP': '_nJbFYL1BEeeOa-zgrqJz2A', 'IR_PUMP': '_CUmMUL05EeeIAZwz_ph7dA', 'TRWS': '_ogj3QL0cEeeZTKIXWf92aQ', 'WM': '_LsmJkL1REeesN7g4ZQQ67w', 'LS': '_-eBakMUTEeeWn40RWAX57w', 'LIFT': '_h-dPkMLXEeeo0tlDmcMRMA', 'TRAV_ESC_1': '_0F_-YMLEEee62KrShjok0Q', 'TRAV_ESC_2': '_B654sMLFEee62KrShjok0Q', 'DIAG_RIO': '_-stmAL1AEeeIAZwz_ph7dA', 'DIAG_NWSW': '_Ml8HAL0-EeeOa-zgrqJz2A','DIAG_PLC': '_SlpW8L_AEeeH--JTscxR2w','DIAG_SER': '_JUeNoL3rEeePMsJSZ95T8Q', 'CPM': '__nPLMLroEeeLy8868mfXIA','CHWP': '_VJ7RUMLQEeeL65l03tuNYA', 'PMU': '_CZJQsMLBEee62KrShjok0Q', 'CDS': '_1f8EcMLSEeetjoi2WK0QMQ', 'CHWH': '_7UVW4MLQEeeL65l03tuNYA', 'HEX': '_UD3Q8L0ZEeeZTKIXWf92aQ', 'PEU': '_acS9ML0aEeeZTKIXWf92aQ', 'VRV': '_gcOKYL09EeeIAZwz_ph7dA', 'FCU': '_rX0coL06EeeIAZwz_ph7dA', 'CRAC': '_3_UL0L0-EeeIAZwz_ph7dA', 'ACS_1': '__FX5cL-4EeeH--JTscxR2w', 'ACS_2': '_HmctMM5gEees1r21d7RbDQ', 'ACMODAMP': '_7a5P8L1AEeeOa-zgrqJz2A'}
eqpt_list_PHE = ['FTC_PUMP', 'SW_PUMP', 'CW_PUMP', 'IR_PUMP', 'LS', 'TRWS', 'WM']
eqpt_list_VHTS = ['LIFT', 'TRAV_ESC']
eqpt_list_BACS = ['DIAG_SER', 'DIAG_PLC', 'DIAG_NWSW', 'DIAG_RIO']
eqpt_list_CHILLER = ['CPM', 'HEX', 'CHWP', 'PMU', 'CDS']
eqpt_list_CHWH = ['CHWH']
eqpt_list_HVAC = ['PEU', 'VRV', 'FCU', 'CRAC']
svg_list = ['VHTS', 'BACS', 'PHE', 'CHILLER', 'CHWH', 'HVAC']
sys_list = [eqpt_list_VHTS, eqpt_list_PHE, eqpt_list_BACS, eqpt_list_CHILLER, eqpt_list_CHWH, eqpt_list_HVAC]
ws_name = '2-Instance List'

col_location = 1
col_sub_loc = 2
col_class_name = 3
col_eqp_name = 4
col_unq_id = 8
wb_org = openpyxl.load_workbook(file_sour)
ws_org = wb_org[ws_name]
wb_new = openpyxl.Workbook()
t_row = ws_org.max_row

for each_svg in svg_list:
    row_list = []
    ws_list = []
    ws_new = wb_new.create_sheet(each_svg)
    for each_rows in ws_org.rows:
        eqpt_list = 'eqpt_list_' + each_svg

        if not each_rows[col_class_name].value:
            continue

        if each_rows[col_class_name].value in sys_list[dict_sys[eqpt_list]]:
            # print(each_rows[col_class_name].value)
            if each_rows[col_class_name].value == 'TRAV_ESC':
                if 'TRAV' in each_rows[col_eqp_name].value or 'Trav' in each_rows[col_eqp_name].value:
                    each_rows[col_class_name].value = each_rows[col_class_name].value + '_2'
                else:
                    each_rows[col_class_name].value = each_rows[col_class_name].value + '_1'
            db_path = ':' + each_rows[col_location].value + ':' + each_rows[col_sub_loc].value + ':' + each_rows[
                col_eqp_name].value
            row_list.append(dict_eq[each_rows[col_class_name].value])
            row_list.append(220)
            row_list.append(735)
            row_list.append(db_path)
            row_list.append(dict_id[each_rows[col_class_name].value])
            row_list.append(each_rows[col_unq_id].value)
            row_list.append(dict_rd[each_rows[col_class_name].value])
            # print(row_list)
            ws_list.append(row_list)
            # print(ws_list)
            row_list = []
            print(ws_list)

    for each in ws_list:
        ws_new.append(each)
wb_new.save(file_new_sour)




