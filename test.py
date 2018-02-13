import openpyxl
import os

cur_dir = os.getcwd()
os.chdir(cur_dir)
# sta_list = ['RNST050', 'RSST010', 'RSST030', 'RSST040',	'RSST070', 'RNST010', 'RNST030', 'RNST040',	'RNST060', 'RNST070', 'RNST020', 'RNST090',	'RNST080', 'GSST010', 'RSST020', 'RSST050', 'RSST060', 'UCST000', 'REST020']
sta_list = []
file_path = cur_dir
file_name = 'image_gen_long.xlsx'
sys_name_list = ['VHTS', 'BACS', 'PHE']
# sys_name = 'VHTS'
# sys_name = 'BACS'
# sys_name = 'PHE'

file_sour = file_path + '\\' + file_name
file_n_sour = file_sour.replace('.xlsx', '_update1.xlsx')
wb = openpyxl.load_workbook(file_sour)
ws = wb['PHE']

ws['H1'].value = ws['D1'].value[5:12]


def upd_col_h(x):
    # 更新h列
    try:
        if ws['H' + str(x - 1)].value[-2:-1] == '_':
            ws['H' + str(x)].value = ws['H' + str(x - 1)].value
        return ws['H' + str(x)].value
    except TypeError:
        print('第%d行有错' %x)


def add_col_h(x):
    # 增加一列，为了判断是否需要增加图
    z = 0
    s1 = ws['H' + str(x - 1)].value[-2:]
    if s1[0] == '_':
        z = int(s1[1]) + 1
    else:
        z = 1
    ws['H' + str(x)].value = ws['D' + str(x)].value[5:12] + '_' + str(z)
    return ws['H' + str(x)].value


total_row = ws.max_row + 1

for i in range(2, total_row):
    # 同一个车站，设备类型类相同
    if ws['A' + str(i)].value == ws['A' + str(i - 1)].value and ws['D' + str(i)].value[0:11] == ws['D' + str(
            i - 1)].value[0:11]:
        if ws['B' + str(i - 1)].value + 550 < 6821 and ws['C' + str(i - 1)].value < 2836:
            ws['B' + str(i)].value = ws['B' + str(i - 1)].value + 550
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)

        elif ws['B' + str(i - 1)].value + 550 > 6821 and ws['C' + str(i - 1)].value < 2536:
            ws['B' + str(i)].value = 220
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value + 300
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)

        elif ws['B' + str(i - 1)].value + 550 < 6821 and ws['C' + str(i - 1)].value > 2834:
            ws['B' + str(i)].value = 220
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)
            ws['C' + str(i)].value = 735

        elif ws['B' + str(i - 1)].value + 550 > 6821 and ws['C' + str(i - 1)].value > 2834:
            ws['B' + str(i)].value = 220
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)
            ws['C' + str(i)].value = 735

    # 同一个车站，但设备类型不同
    elif ws['A' + str(i)].value != ws['A' + str(i - 1)].value and ws['D' + str(i)].value[0:11] == ws['D' + str(
            i - 1)].value[0:11]:
        if ws['C' + str(i - 1)].value < 2536:
            ws['B' + str(i)].value = 220
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value + 300
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)
        else:
            ws['B' + str(i)].value = 220
            ws['C' + str(i)].value = 735
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)

    # 不同车站
    elif ws['D' + str(i)].value[0:11] != ws['D' + str(i - 1)].value[0:11]:
        ws['B' + str(i)].value = 220
        ws['C' + str(i)].value = 735
        ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]

    elif ws['C' + str(i - 1)].value > 2834:
        # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
        ws['H' + str(i)].value = add_col_h(i)
        ws['C' + str(i)].value = 735

    elif ws['A' + str(i)].value == '':
        ws['H' + str(i)].value = ws['H' + str(i - 1)].value

wb.save(file_n_sour)