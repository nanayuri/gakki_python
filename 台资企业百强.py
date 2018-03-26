import openpyxl
import re

file = '123.txt'
company_list = []
with open(file) as f:
    content = f.readlines()
print(content)
company_year = ''
company_boss = ''
for each_line in content:
    if each_line.find('年，') != -1:
        company_year = str(each_line.split('年，')[0])[-4:]
    if each_line.find('年。') != -1:
        company_year = str(each_line.split('年。')[0])[-4:]
    if each_line.find('成立于') != -1:
        company_year = str(each_line.split('成立于')[1])[0:4]
    if each_line.find('年创办，') != -1:
        company_year = str(each_line.split('年创办，')[0])[-4:]
    if each_line.find('年由') != -1:
        company_year = str(each_line.split('年由')[0])[-4:]
    if each_line.find('年创建，') != -1:
        company_year = str(each_line.split('年创建，')[0])[-4:]
    if each_line.find('董事长') != -1:
        company_boss = str(each_line.split('董事长')[1])[0:3]
    if each_line.find('创立于') != -1:
        company_boss = str(each_line.split('创立于')[0])[-3:]
    if each_line.find('于19') != -1:
        if str(each_line.split('于19')[0])[-3:].find('创立于') != -1:
            company_boss = str(each_line.split('于19')[0])[-3:]

    if each_line.find('创办于') != -1:
        if str(each_line.split('创办于')[0])[-3].find(':') != -1:
            company_boss = str(each_line.split('创办于')[0])[-4]
    if each_line.find('公司：') != -1:
        company_name = str(each_line.split('：')[0]).replace('\u3000\u3000','') + '\n'
        company_n = [company_name, company_year, company_boss]
        company_list.append(company_n)
    if company_year != '':
        company_year != ''
    if company_boss != '':
        company_boss != ''
print(company_list)
wb = openpyxl.Workbook()
wb.guess_types = True
ws = wb.create_sheet('company')
head = ['公司名称', '成立年份', '老板']
ws.append(head)
for each in company_list:
    ws.append(each)
'''
for i in range(len(company_list)):
    print(i)
    ws['A' + str(i + 2)].value = company_list[i]
'''
wb.save('台资企业百强名单.xlsx')
wb.close()
