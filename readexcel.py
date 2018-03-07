# Project: Doha
# Written by: Huang Chao
# Date: 2018-02-13
# .svg image files auto generation

import openpyxl
import os

cur_dir = os.getcwd()
os.chdir(cur_dir)
# sta_list = ['RNST050', 'RSST010', 'RSST030', 'RSST040',	'RSST070', 'RNST010', 'RNST030', 'RNST040',	'RNST060', 'RNST070', 'RNST020', 'RNST090',	'RNST080', 'GSST010', 'RSST020', 'RSST050', 'RSST060', 'UCST000', 'REST020']
sta_list = []
file_path = cur_dir
file_name = 'image_autogen.xlsx'
#sys_name_list = ['VHTS', 'BACS', 'PHE', 'CHILLER']
sys_name_list = ['HVAC']
# sys_name = 'VHTS'
# sys_name = 'BACS'
# sys_name = 'PHE'

file_sour = file_path + '\\' + file_name
file_n_sour = file_sour.replace('.xlsx', '_update.xlsx')
station_name = ''
lines = []
long_str = ''
a = r'  <use mc:refType="symbol" mc:navigationId="" mc:layerName="'
b1 = r'" mc:navImageExtent="" transform="matrix(1 0 0 1 '
b2 = r')" mc:dbPath="'
c = r'" xlink:show="embed" xlink:type="simple" mc:navImageId="" mc:protoId="'
d = r'" mc:hvEntityId="'
e = r'" xlink:href="#'
f = r'" xlink:actuate="onLoad"/>'
a1_er = r'  <use mc:refType="symbol" mc:layerName="default" mc:navigationId="'
a2_er = r'  <use mc:refType="symbol" mc:layerName="default" mc:navigationId="'
b1_er = r'" mc:navImageExtent="" transform="matrix(4 0 0 4 2750 100)" mc:dbPath="NAV_ARROW1" xlink:show="embed" xlink:type="simple" mc:protoId="_cZbP8OUzEeWqL4Ot-6sQ7Q_1457444481647" mc:navImageId="" mc:hvEntityId="NAV_ARROW1" xlink:href="#_V7EnMETnEeWaHvxsOBt1WQ" xlink:actuate="onLoad"/>'
b2_er = r'" mc:navImageExtent="" transform="matrix(4 0 0 4 4500 100)" mc:dbPath="NAV_ARROW2" xlink:show="embed" xlink:type="simple" mc:protoId="_cZbP8OUzEeWqL4Ot-6sQ7Q_1457444481647" mc:navImageId="" mc:hvEntityId="NAV_ARROW2" xlink:href="#_LirekES8EeWaHvxsOBt1WQ" xlink:actuate="onLoad"/>'

image_cat = ['BACS', 'PHE', 'VHTS', 'CHILLER']
dict2 = {'BACS': 'Bacs', 'PHE': 'Phe', 'VHTS': 'VHTS', 'CHILLER': 'Chiller', 'HVAC': 'HVAC'}
out_des = 'Y:\\git_depots\\cfg\\occ\\model\\metaconf\\proj\\ImageProject_ISCS\\svg-images\\DOH\\STA\\Bacs\\DOH\\STA\\'
dict3 = {'BACS_UCST000_CHWP_TUN_CHWP_1': 'BACS_UCST000_CHWH_CHW_HD_TUN', 'BACS_UCST000_CHWP_TUN_CHWP_2': 'BACS_UCST000_CHWH_CHW_HD_TUN', 'BACS_UCST000_CHWP_TUN_CHWP_3': 'BACS_UCST000_CHWH_CHW_HD_TUN', 'BACS_UCST000_CHWP_STN_CHWP_1': 'BACS_UCST000_CHWH_CHW_HD_PL', 'BACS_UCST000_CHWP_STN_CHWP_2': 'BACS_UCST000_CHWH_CHW_HD_PL', 'BACS_UCST000_CHWP_STN_CHWP_3': 'BACS_UCST000_CHWH_CHW_HD_PL', 'BACS_RSST070_CHWP_CWSP_AA_01': 'BACS_RSST070_CHWH_CHW_CHI_AA_01', 'BACS_RSST070_CHWP_CWSP_AA_02': 'BACS_RSST070_CHWH_CHW_CHI_AA_02', 'BACS_RSST060_CHWP_CWSP_AA_01': 'BACS_RSST060_CHWH_CHW_01', 'BACS_RSST060_CHWP_CWSP_DC_AA_01': 'BACS_RSST060_CHWH_CHW_01', 'BACS_RSST050_CHWP_CWSP_AA_01': 'BACS_RSST050_CHWH_CHW_CHI_AA_01', 'BACS_RSST050_CHWP_CWSP_AA_02': 'BACS_RSST050_CHWH_CHW_CHI_AA_02', 'BACS_RNST010_CHWP_CWSP_AA_01': 'BACS_RNST010_CHWH_CHW_01', 'BACS_RNST030_CHWP_CWSP_AA_01': 'BACS_RNST030_CHWH_CHW_01', 'BACS_RNST020_CHWP_CWSP_AA_01': 'BACS_RNST020_CHWH_CHW_01', 'BACS_RNST040_CHWP_WCP_01': 'BACS_RNST040_CHWH_CHW_01', 'BACS_RNST050_CHWP_CWSP_AA_01': 'BACS_RNST050_CHWH_CHW_01', 'BACS_RNST050_CHWP_CWSP_AA_02': 'BACS_RNST050_CHWH_CHW_02', 'BACS_RNST050_CHWP_CWSP_AA_03': 'BACS_RNST050_CHWH_CHW_03', 'BACS_RNST060_CHWP_CWSP_AA_03': 'BACS_RNST060_CHWH_CHW_01', 'BACS_RNST070_CHWP_CWSP_AA_01': 'BACS_RNST070_CHWH_CHW_01'}
dict4 = {}
dict_id = {'FTC_PUMP': 'BasicEquipment_DohBacsFtcPump_EVqHEL3nEeeHE768d9D71g_1509418578433',  'SW_PUMP': 'BasicEquipment_DohBacsSwPump_8fpicL1JEeekaonQy-nY5g_1509351094807', 'CW_PUMP': 'BasicEquipment_DohBacsCwPump_3jnqYLo1Eee4FZNjhxyf2w_1509012618246', 'IR_PUMP': 'BasicEquipment_DohBacsIrPump_HNYowLrfEeeUOKN4pirLLA_1509085308236', 'TRWS': 'BasicEquipment_DohBacsTrws_rn29YLrBEee2yfXdVkK8vg_1509072667702', 'WM': 'BasicEquipment_DohBacsWm_ZhZJcL1BEeeMG7pPixeMOg_1509347424135', 'LS': 'BasicEquipment_DohBacsLs_9GgFcMT8Eee2FLSNuBaW1Q_1510197636935', 'LIFT': 'BasicEquipment_DohBacsLift_3yJJ4MLNEeetjoi2WK0QMQ_1509957512574', 'TRAV_ESC_1': 'BasicEquipment_DohBacsTravEsc_FZrJ0MK-Eee62KrShjok0Q_1509950732013', 'TRAV_ESC_2': 'BasicEquipment_DohBacsTravEsc_FZrJ0MK-Eee62KrShjok0Q_1509950732013', 'DIAG_RIO': 'BasicEquipment_DohBacsDiagRio_BSA2gLrfEeeUOKN4pirLLA_1509085268456', 'DIAG_NWSW': 'BasicEquipment_DohBacsDiagNwsw_1bxNELo1Eee4FZNjhxyf2w_1509012604001','DIAG_PLC': 'BasicEquipment_DohBacsDiagPlc_66ly8L8nEee5TNQefPaSHQ_1509556383647','DIAG_SER': 'BasicEquipment_DohBacsDiagSer_78GJcL1JEeekaonQy-nY5g_1509351091079', 'CPM': 'BasicEquipment_DohBacsCpm_vyjcMLrBEee2yfXdVkK8vg_1509072695667', 'CHWP': 'BasicEquipment_DohBacsChwp_XUuHMMLLEeeL65l03tuNYA_1509956435748', 'PMU': 'BasicEquipment_DohBacsPmu_Hx6MgMK-Eee62KrShjok0Q_1509950747976', 'CDS': 'BasicEquipment_DohBacsCds_5O-7UMLNEeetjoi2WK0QMQ_1509957522309', 'CHWH': 'BasicEquipment_DohBacsChwh_Ysi1wMLLEeeL65l03tuNYA_1509956444956', 'HEX': 'BasicEquipment_DohBacsHex_yb5dALrBEee2yfXdVkK8vg_1509072713424', 'PEU': 'BasicEquipment_DohBacsPeu_tv7dILrBEee2yfXdVkK8vg_1509072681971', 'VRV': 'BasicEquipment_DohBacsVrv_FTr8MLrfEeeUOKN4pirLLA_1509085295475', 'FCU': 'BasicEquipment_DohBacsFcu_GOmsULrfEeeUOKN4pirLLA_1509085301653', 'CRAC': 'BasicEquipment_DohBacsCrac_ETayALrfEeeUOKN4pirLLA_1509085288736', 'ACS_1': 'BasicEquipment_DohBacsAcs_FY_MUL3nEeeHE768d9D71g_1509418585493', 'ACS_2': 'BasicEquipment_DohBacsAcs_FY_MUL3nEeeHE768d9D71g_1509418585493', 'ACMODAMP': 'BasicEquipment_DohBacsAcModamp_6U7UYLo1Eee4FZNjhxyf2w_1509012636838'}
dict_rd = {'FTC_PUMP': '_yAuL0L_CEeeok6BbH_neNQ',  'SW_PUMP': '_jy7a4L4WEeeBbKd0rZ4ZWQ', 'CW_PUMP': '_nJbFYL1BEeeOa-zgrqJz2A', 'IR_PUMP': '_CUmMUL05EeeIAZwz_ph7dA', 'TRWS': '_ogj3QL0cEeeZTKIXWf92aQ', 'WM': '_LsmJkL1REeesN7g4ZQQ67w', 'LS': '_-eBakMUTEeeWn40RWAX57w', 'LIFT': '_h-dPkMLXEeeo0tlDmcMRMA', 'TRAV_ESC_1': '_0F_-YMLEEee62KrShjok0Q', 'TRAV_ESC_2': '_B654sMLFEee62KrShjok0Q', 'DIAG_RIO': '_-stmAL1AEeeIAZwz_ph7dA', 'DIAG_NWSW': '_Ml8HAL0-EeeOa-zgrqJz2A','DIAG_PLC': '_SlpW8L_AEeeH--JTscxR2w','DIAG_SER': '_JUeNoL3rEeePMsJSZ95T8Q', 'CPM': '__nPLMLroEeeLy8868mfXIA','CHWP': '_VJ7RUMLQEeeL65l03tuNYA', 'PMU': '_CZJQsMLBEee62KrShjok0Q', 'CDS': '_1f8EcMLSEeetjoi2WK0QMQ', 'CHWH': '_7UVW4MLQEeeL65l03tuNYA', 'HEX': '_UD3Q8L0ZEeeZTKIXWf92aQ', 'PEU': '_acS9ML0aEeeZTKIXWf92aQ', 'VRV': '_gcOKYL09EeeIAZwz_ph7dA', 'FCU': '_rX0coL06EeeIAZwz_ph7dA', 'CRAC': '_3_UL0L0-EeeIAZwz_ph7dA', 'ACS_1': '__FX5cL-4EeeH--JTscxR2w', 'ACS_2': '_HmctMM5gEees1r21d7RbDQ', 'ACMODAMP': '_7a5P8L1AEeeOa-zgrqJz2A'}


def upd_col_h(x):
    # 更新h列
    try:
        if ws['H' + str(x - 1)].value[-2:-1] == '_':
            ws['H' + str(x)].value = ws['H' + str(x - 1)].value
        return ws['H' + str(x)].value
    except TypeError:
        print('第%d行有错' %x)


def add_col_h(x):
    # 增加一列，为了判断是否需要增加图
    z = 0
    s1 = ws['H' + str(x - 1)].value[-2:]
    if s1[0] == '_':
        z = int(s1[1]) + 1
    else:
        z = 1
    ws['H' + str(x)].value = ws['D' + str(x)].value[5:12] + '_' + str(z)
    return ws['H' + str(x)].value


def ck_unq_sta(x, y):
    arr_cat_prv = ''
    arr_cat_fwd = ''
    nav1_id = ''
    nav2_id = ''
    if len(y) == 9:
        if x == 'BACS' and y + '_' + x in sys_sta_list:
            arr_cat_prv = x
            arr_cat_fwd = ''
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                      dict2[x] + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er
        elif x == 'PHE' and y + '_' + x in sys_sta_list:
            arr_cat_prv = x
            arr_cat_fwd = 'BACS'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                      dict2[x] + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                      0:7] + '_' + \
                      dict2[arr_cat_fwd] + '_Summary' + b2_er
        elif x == 'VHTS' and y + '_' + x in sys_sta_list:
            arr_cat_prv = x
            arr_cat_fwd = 'PHE'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                      dict2[x] + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                      0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er
        elif x == 'CHILLER' and y + '_' + x in sys_sta_list:
            arr_cat_prv = x
            arr_cat_fwd = 'Electrical'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                      dict2[x] + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                    0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er
        elif x == 'HVAC' and y + '_' + x in sys_sta_list:
            for each_index in range(1, 10):
                if '_' + str(each_index) in y:
                    if y.replace(('_' + str(each_index)), ('_' + str(each_index + 1))) + '_' + x in sys_sta_list:
                        arr_cat_prv = x
                        arr_cat_fwd = x
                        if each_index == 1:
                            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                                      dict2[x] + '_Summary' + b1_er
                            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[arr_cat_fwd] + '_' + str(each_index + 1) + '_Summary' + b2_er
                        else:
                            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                                      dict2[x] + '_' + str(each_index - 1) + '_Summary' + b1_er
                            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                        0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                  0:7] + '_' + \
                                      dict2[arr_cat_fwd] + '_' + str(
                                each_index + 1) + '_Summary' + b2_er
                    else:
                        arr_cat_prv = x
                        arr_cat_fwd = 'CHILLER'
                        if each_index == 1:
                            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                                      dict2[x] + '_Summary' + b1_er
                            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                        0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                  0:7] + '_' + dict2[arr_cat_fwd] + '_Summary' + b2_er
                        else:
                            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + \
                                      dict2[x] + '_' + str(each_index - 1) + '_Summary' + b1_er
                            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                        0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                  0:7] + '_' + \
                                      dict2[arr_cat_fwd] + '_Summary' + b2_er
                    break

    elif len(y) == 7 and y + '_1_' + x in sys_sta_list:
        if x == 'VHTS':
            arr_cat_prv = 'Electrical'
            arr_cat_fwd = x
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                      dict2[x] + '_Summary' + b2_er
        elif x == 'PHE':
            if y + '_1_' + 'VHTS' in sys_sta_list:
                arr_cat_prv = 'VHTS'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_1_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                    x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                          dict2[x] + '_Summary' + b2_er
            else:
                arr_cat_prv = 'VHTS'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                    x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                          dict2[x] + '_Summary' + b2_er
        elif x == 'BACS':
            if y + '_1_' + 'PHE' in sys_sta_list:
                arr_cat_prv = 'Phe'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_1_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                    x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                          dict2[x] + '_Summary' + b2_er
            else:
                arr_cat_prv = 'Phe'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                    x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                          dict2[x] + '_Summary' + b2_er
        elif x == 'CHILLER':
            arr_cat_prv = 'HVAC'
            arr_cat_fwd = x
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                      dict2[x] + '_Summary' + b2_er
        elif x == 'HVAC':
            arr_cat_prv = 'SVS'
            arr_cat_fwd = x
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[0:7] + '_' + dict2[
                x] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y + '_1_' + \
                      dict2[x] + '_Summary' + b2_er

    else:
        if x == 'VHTS':
            arr_cat_prv = 'Electrical'
            arr_cat_fwd = 'Phe'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er
        elif x == 'PHE':
            if y + '_1_' + 'VHTS' in sys_sta_list:
                arr_cat_prv = 'VHTS'
                arr_cat_fwd = 'Bacs'
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_1_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_fwd + '_Summary' + b2_er

            else:
                arr_cat_prv = 'VHTS'
                arr_cat_fwd = 'Bacs'
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_fwd + '_Summary' + b2_er
        elif x == 'BACS':
            if y + '_1_' + 'PHE' in sys_sta_list:
                arr_cat_prv = 'Phe'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_1_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          dict2[arr_cat_fwd] + '_Summary' + b2_er
            else:
                arr_cat_prv = 'Phe'
                arr_cat_fwd = x
                nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          arr_cat_prv + '_Summary' + b1_er
                nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                             0:7] + '_' + dict2[arr_cat_fwd] + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                        0:7] + '_' + \
                          dict2[arr_cat_fwd] + '_Summary' + b2_er
        elif x == 'CHILLER':
            arr_cat_prv = 'HVAC'
            arr_cat_fwd = 'Electrical'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er
        elif x == 'HVAC':
            arr_cat_prv = 'SVS'
            arr_cat_fwd = 'Chiller'
            nav1_id = a1_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_prv + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_prv + '_Summary' + b1_er
            nav2_id = a2_er + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                       0:7] + '_' + arr_cat_fwd + '_' + 'DOH_STA_Bacs_DOH_STA_' + y[
                                                                                                                  0:7] + '_' + \
                      arr_cat_fwd + '_Summary' + b2_er

    return arr_cat_prv, arr_cat_fwd, nav1_id, nav2_id


def upd_CHWP_cord(x):
    return ws['B' + str(x)].value - 100


def insert_row(x):
    ws.insert_rows(x + 1, 1)
    ws['A' + str(x + 1)].value = 'DohBacsChwh_layer'
    ws['B' + str(x + 1)].value = ws['B' + str(x)].value + 200
    ws['C' + str(x + 1)].value = ws['C' + str(x)].value
    ws['D' + str(x + 1)].value = dict4[dict3[ws['F' + str(x)].value]]
    ws['E' + str(x + 1)].value = dict_id['CHWH']
    ws['F' + str(x + 1)].value = dict3[ws['F' + str(x)].value]
    ws['G' + str(x + 1)].value = dict_rd['CHWH']
    ws['H' + str(x + 1)].value = ws['H' + str(x)].value


def to_svg(data):
    # output to local
    with open(f_out_dir + '\\' + 'Summary_Background_' + sta_code + '_' + each_cat + '.svg', 'w') as f_temp_final:
        # output to VM cfg repository
        # with open(temp_vm, 'w') as f_temp_final:
        # print(temp_vm_fn)
        content1 = f_temp1.read()[0:-7]
        content1 = content1.replace('Hamad International Airport', dict1[sta_code[0:7]])
        content1 = content1.replace('REST020', sta_code[0:7])
        content2 = '\n' + '</svg>'
        f_temp_final.write(content1 + data + content2)
        os.chdir(cur_dir)


def calc_cord(i):
    origin_x = 220
    origin_y = 735
    upper_x = 6820
    upper_y = 2835
    x_step = 550
    y_step = 300
    if ws['A' + str(i)].value == ws['A' + str(i - 1)].value and ws['D' + str(i)].value[0:11] == ws['D' + str(
            i - 1)].value[0:11]:
        if ws['B' + str(i - 1)].value + x_step < upper_x + 1 and ws['C' + str(i - 1)].value < upper_y + 1:
            ws['B' + str(i)].value = ws['B' + str(i - 1)].value + x_step
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)

        elif ws['B' + str(i - 1)].value + x_step > upper_x + 1 and ws['C' + str(i - 1)].value < upper_y + 1 - y_step:
            ws['B' + str(i)].value = origin_x
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value + y_step
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)

        elif ws['B' + str(i - 1)].value + x_step < upper_x + 1 and ws['C' + str(i - 1)].value > upper_y - 1:
            ws['B' + str(i)].value = origin_x
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)
            ws['C' + str(i)].value = origin_y

        elif ws['B' + str(i - 1)].value + x_step > upper_x + 1 and ws['C' + str(i - 1)].value > upper_y - 1:
            ws['B' + str(i)].value = origin_x
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)
            ws['C' + str(i)].value = origin_y

    # 同一个车站，但设备类型不同
    elif ws['A' + str(i)].value != ws['A' + str(i - 1)].value and ws['D' + str(i)].value[0:11] == ws['D' + str(
            i - 1)].value[0:11]:
        if ws['C' + str(i - 1)].value < upper_y + 1 - y_step:
            ws['B' + str(i)].value = origin_x
            ws['C' + str(i)].value = ws['C' + str(i - 1)].value + y_step
            ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]
            ws['H' + str(i)].value = upd_col_h(i)
        else:
            ws['B' + str(i)].value = origin_x
            ws['C' + str(i)].value = origin_y
            # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
            ws['H' + str(i)].value = add_col_h(i)

    # 不同车站
    elif ws['D' + str(i)].value[0:11] != ws['D' + str(i - 1)].value[0:11]:
        ws['B' + str(i)].value = origin_x
        ws['C' + str(i)].value = origin_y
        ws['H' + str(i)].value = ws['D' + str(i)].value[5:12]

    elif ws['C' + str(i - 1)].value > upper_y - 1:
        # ws['H' + str(i)].value = ws['D' + str(i)].value[5:12] + '_1'
        ws['H' + str(i)].value = add_col_h(i)
        ws['C' + str(i)].value = origin_y

    elif ws['A' + str(i)].value == '':
        ws['H' + str(i)].value = ws['H' + str(i - 1)].value

    if ws['H' + str(i)].value not in sta_list:
        sta_list.append(ws['H' + str(i)].value)

sys_sta_list = []
wb = openpyxl.load_workbook(file_sour)
ws_chwh = wb['CHWH']
for i in range(1, ws_chwh.max_row + 1):
    dict4[ws_chwh['F' + str(i)].value] = ws_chwh['D' + str(i)].value

for sys_name in sys_name_list:
    sta_list = []
    image_name = '_' + sys_name + '_Auto_gen'
    if not os.path.exists(cur_dir + '\\' + sys_name):
        os.mkdir(sys_name)


    ws = wb[sys_name]

    total_row = ws.max_row + 1

    ws['H1'].value = ws['D1'].value[5:12]

    for i in range(2, total_row):
        calc_cord(i)

    for j in range(1, 1000):
        if ws['A' + str(j)].value == 'DohBacsChwp_layer' and ws['F' + str(j)].value in dict3:
            ws['B' + str(j)].value = upd_CHWP_cord(j)
            insert_row(j)

    os.chdir(cur_dir + '\\' + sys_name)

    for each_file in os.listdir(os.getcwd()):
        os.remove(each_file)

    for each_sta in sta_list:
        for each_rows in ws.rows:
            if each_rows[7].value == each_sta:
                long_str += a + each_rows[0].value + b1 + str(each_rows[1].value) + ' ' + str(each_rows[2].value) + b2 + each_rows[3].value + c + each_rows[4].value + d + each_rows[5].value + e + each_rows[6].value + r'" xlink:actuate="onLoad"/>' + '\n'

        with open(str(each_sta) + image_name + '.svg', 'w') as f:
            f.write(long_str)
        long_str = ''

        if str(each_sta) + '_' + sys_name not in sys_sta_list and os.path.getsize(str(each_sta) + image_name + '.svg') != 0:
            sys_sta_list.append(str(each_sta) + '_' + sys_name)
    os.chdir(cur_dir)
#print(sta_list)
print(sys_sta_list)
wb.save(file_n_sour)
os.chdir(cur_dir)

if not os.path.exists(cur_dir + '\\' + 'output_files'):
    os.mkdir('output_files')
f_out_dir = cur_dir + '\\' + 'output_files'
os.chdir(f_out_dir)
for each_file in os.listdir(cur_dir + '\\' + 'output_files'):
    os.remove(each_file)

for each_cat in sys_name_list:
    sour_dir = cur_dir + '\\' + each_cat
    temp_file1 = cur_dir + '\\Template\\Summary_Background_template_' + each_cat + '_1.svg'
    temp_file2 = cur_dir + '\\Template\\Summary_Background_template_' + each_cat + '_2.svg'
    list1 = list(os.walk(sour_dir))[0][2]
    str_add = ''
    dict1 = {'RNST050': 'Al Qassar', 'RSST010': 'Al Doha Al Jadeeda', 'RSST030': 'Al Matar', 'RSST040': 'Oqba Ibn Nafie', 'RSST070': 'AL Wakra', 'RNST010': 'Al Bidda', 'RNST030': 'Corniche', 'RNST040': 'Doha Exhibition and Convention Centre', 'RNST060': 'Katara', 'RNST070': 'Legtaifiya', 'RNST020': 'West Bay', 'RNST090': 'Lusail', 'RNST080': 'Qatar University', 'GSST010': 'Al Mansoura', 'RSST020': 'Umm Ghuwailina', 	'RSST050': 'Economic Zone', 'RSST060': 'Ras Bu Fontas', 'UCST000': 'Msheireb', 'REST020': 'Hamad International Airport'}

    for each in list1:
        each = sour_dir + '\\' + each
        file_name = os.path.basename(each)
        for each_num in range(1, 11):
            if file_name.find('_' + str(each_num)) is not -1:
                sta_code = file_name[0:9]
                break
            else:
                sta_code = file_name[0:7]
        str_line = ''
        temp_vm_fn = '\\DOH_STA_Bacs_DOH_STA_' + sta_code + '_' + dict2[each_cat] + '_Summary.svg'
        temp_vm = out_des + sta_code[0:7] + '\\' + dict2[each_cat] + temp_vm_fn

        with open(each) as f_sour:
            lines = f_sour.readlines()
            try:
                if lines[-1].find('dbPath=""') != -1:
                    lines = lines[0:-1]
            except IndexError:
                print(each)

            for each_line in lines:
                str_line += each_line

            arr = ck_unq_sta(each_cat, sta_code)
            print(each_cat, sta_code)
            # print(arr)

            if each_cat == 'BACS':
                if sta_code + '_1_' + each_cat in sys_sta_list:
                    str_line += arr[2]
                    str_line += '\n'
                    str_line += arr[3]
                else:
                    str_line += arr[2]
            else:
                str_line += arr[2]
                str_line += '\n'
                str_line += arr[3]

            with open(temp_file1) as f_temp1, open(temp_file2) as f_temp2:
                os.chdir(cur_dir)
                to_svg(str_line)
                os.chdir(cur_dir)
