import openpyxl
import time
from openpyxl import load_workbook
import os
import csv

Class_list_dict = {}
Class_list_dict['isAComment'] = 1
Class_list_dict['Class name'] = 4
Class_list_dict['CSS'] = 5
Class_list_dict['type'] = 10
Class_list_dict['bit_0'] = 15
Class_list_dict['bit_1'] = 16
Class_list_dict['bit_2'] = 17
Class_list_dict['bit_3'] = 18
Class_list_dict['bit_default'] = 19
Class_list_dict['SIL Level'] = 11
Class_list_dict['code'] = 63
Instance_list_dict = {}
Instance_list_dict['isAComment'] = 1
Instance_list_dict['Description'] = 5
Instance_list_dict['Location'] = 6
Instance_list_dict['Sublocation'] = 7
Instance_list_dict['Class name'] = 8
Instance_list_dict['Equipment name'] = 9
Instance_list_dict['RTU name'] = 10
Instance_list_dict['Short name'] = 20
Instance_list_dict['Unique ID code'] = 21
EV_dict = {}
EV_dict['Class name'] = 0
EV_dict['iopath'] = 1
EV_dict['vetype'] = 2
EV_dict['ve_base'] = 3
EV_dict['description'] = 4
EV_dict['address'] = 5
EV_dict['evgroup'] = 6
EV_dict['deadband'] = 8
EV_dict['length'] = 12

type_dict = {
    'dci': 'dac',
    'aci': 'aac',
    'dio': 'dov',
    'aio': 'aio'
}

ev_type_dict = {
    'dci': 'dei',
    'aci': 'aei',
    'dio': 'deo',
    'aio': 'aeo'
}

sil_type_dict = {
    'SIL0': '_S0',
    'SIL2_Monitoring': '_S2M',
    'SIL2_Control': '_S2C'
}

ins_list = []
list_nom = [[], [], [], [], [], [], [], [], [], []]
ws_list = []
row_list = []
ws_par_list = []
link_list = []
ev_gen_list = []
#label_class_list = ['IDP', 'MSFD', 'VENTILATION', 'ACB', 'ADS', 'EL', 'LBS', 'MCB', 'MCCB', 'MDB', 'MPLCMMS', 'SMDB', 'SPLCMMS', 'TRF']
label_class_list = ['66KVHVMV', '33KVSTMV', '66KVBSMV', '33KVBSMV', '33KVOFMV', '33KVVM', '66KVVM', 'AUXTF', 'POWTF', 'BCA']
all_label = []
all_hvid = []


def save_to_excel(data, i):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(['nom_instance', 'ID', 'ELEMENT_NAME', '__PROTOTYPE__', '__AGGREGATE__', 'link_object_is_from_object_type', 'reference_import', 'label', 'shortname', 'hvid'])
    for each in data:
        ws_new.append(each)
    file_n = str(i) + "-equipment.xlsx"
    wb_new.save(file_n)
    wb_new.close()
    xlsx2csv(file_n)


def save_link_to_excel(data):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(['nom_instance', 'ID', 'ELEMENT_NAME', 'lk0', 'lk1'])
    for each in data:
        ws_new.append(each)
    file_n = "lk_gen.xlsx"
    wb_new.save(file_n)
    wb_new.close()
    xlsx2csv(file_n)


def save_ev_to_excel(data):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(['ID', 'nom_instance', 'reference_import', 'address', 'deadband', 'label', 'length', 'name', 'type', 'ELEMENT_NAME', 'varInvalid1', 'varInvalid2', 'varInvalid3', 'functCheck', 'functTrans'])
    for each in data:
        ws_new.append(each)
    file_n = "ev_gen.xlsx"
    wb_new.save(file_n)
    wb_new.close()
    xlsx2csv(file_n)


def add_parent_path(row_id):
    row_id = row_id[5:]
    id_list = row_id.split(':')
    instance = ':R:A'
    for each_ins in id_list:
        instance += (':' + each_ins)
        if instance not in ins_list:
            ins_list.append(instance)

    return ins_list


def split_parent(ins_list):
    for each_ins in ins_list:
        num = each_ins.count(':')
        if each_ins not in list_nom[num - 1]:
            list_nom[num - 1].append(each_ins)
    return list_nom


def judge_str_null(s):
    if s.strip() == '':
        pass
    else:
        return s


def judge_proto(s_eq, eq):
    if s_eq == '':
        return 'Qatar:QatarAlmSynthe'
    else:
        return eq


def judge_label(s_eq, s_nom):
    if s_eq == '':
        return 'Synthesis class ' + s_nom
    else:
        return 'LusHK:Lus'


def open_excel(file_name):

    wb = openpyxl.load_workbook(file_name)
    ws_1 = wb['1-Class List']
    ws_2 = wb['2-Instance List']
    ws_3 = wb['3-Class Mapping']
    ws_4 = wb['4-External Variables']


def split_parent_node(data):
    par_list = []
    ws_par_list = []
    for each_list in data:
        par_id = each_list
        par_proto = 'Qatar:QatarAlmSynthe'
        par_nom_ins = par_id.split(':')[-1]
        par_elename = ''
        par_aggreg = ''
        par_linkob = ''
        par_refimp = ''
        par_label = 'Synthesis class ' + par_nom_ins
        par_shtnam = '&' + par_nom_ins
        par_hvid = 'STA_' + par_nom_ins
        for each in [par_nom_ins, par_id, par_elename, par_proto, par_aggreg, par_linkob, par_refimp, par_label,
                     par_shtnam, par_hvid]:
            par_list.append(each)
        ws_par_list.append(par_list)
    return ws_par_list


def equipment_csv(file_name):
    # file_name = 'LRT-QSTT-XX-ICD-CCS-GEN00-373543-B- SCADA-DPB-AppendixC_ATT1v0.37_HK_WP7.xlsx'
    wb = openpyxl.load_workbook(file_name, data_only=True)
    list_par = []
    ws_1 = wb['1-Class List']
    ws_2 = wb['2-Instance List']
    ws_3 = wb['3-Class Mapping']
    ws_4 = wb['4-External Variables']
    class_map_dict = {}
    for each_row in ws_3.rows:

        if each_row[0].value != 'Class name':
            class_map_dict[each_row[0].value] = each_row[1].value
    print(class_map_dict)

    for each_row in ws_2.rows:
        row_list = []
        print(each_row[Instance_list_dict['isAComment']].value)
        if each_row[Instance_list_dict['isAComment']].value == 'N':
            row_id = ':R:A:' + judge_str_null(each_row[Instance_list_dict['Location']].value) + ':' + judge_str_null(
                each_row[Instance_list_dict['Sublocation']].value) + ':' + judge_str_null(each_row[Instance_list_dict['Equipment name']].value)
            row_par = ':R:A:' + judge_str_null(each_row[Instance_list_dict['Location']].value) + ':' + judge_str_null(
                each_row[Instance_list_dict['Sublocation']].value)

            try:
                row_proto = class_map_dict[judge_proto(each_row[Instance_list_dict['Equipment name']].value, each_row[Instance_list_dict['Class name']].value)]
            except KeyError:
                row_proto = 'not defined'
            row_nom_ins = row_id.split(':')[-1]
            row_elename = ''
            row_aggreg = ''
            row_linkob = ''
            row_refimp = ''
            row_label = each_row[Instance_list_dict['Description']].value
            row_shtnam = each_row[Instance_list_dict['Short name']].value
            row_hvid = each_row[Instance_list_dict['Unique ID code']].value
            list_par = split_parent(add_parent_path(row_par))
            print(row_id)
            for each in [row_nom_ins, row_id, row_elename, row_proto, row_aggreg, row_linkob, row_refimp, row_label, row_shtnam, row_hvid]:
                row_list.append(each)
            ws_list.append(row_list)

    save_to_excel(ws_list, 'all')
    i = 1
    for each_list in list_par:
        ws_par_list = []
        if each_list:
            for each_data in each_list:
                par_list = []
                par_id = each_data
                par_proto = 'Qatar:QatarAlmSynthe'
                par_nom_ins = par_id.split(':')[-1]
                par_elename = ''
                par_aggreg = ''
                par_linkob = ''
                par_refimp = ''
                par_label = 'Synthesis class ' + par_nom_ins
                par_shtnam = '&' + par_nom_ins
                par_hvid = 'STA_' + par_nom_ins
                for each in [par_nom_ins, par_id, par_elename, par_proto, par_aggreg, par_linkob, par_refimp, par_label,
                             par_shtnam, par_hvid]:
                    par_list.append(each)
                ws_par_list.append(par_list)
            save_to_excel(ws_par_list, i)
            i += 1


def ev(file_name):
    ws_ev_list = []
    class_list = []
    ev_gp_list = []
    # file_name = 'LRT-QSTT-XX-ICD-CCS-GEN00-373543-B- SCADA-DPB-AppendixC_ATT1v0.37_HK_WP7.xlsx'
    wb = openpyxl.load_workbook(file_name, data_only=True)
    list_par = []
    ws_1 = wb['1-Class List']
    ws_2 = wb['2-Instance List']
    ws_3 = wb['3-Class Mapping']
    ws_4 = wb['4-External Variables']
    for each_row in ws_1.rows:

        Class_list_dict['isAComment'] = 1
        Class_list_dict['Class name'] = 4
        Class_list_dict['SIL Level'] = 11
        Class_list_dict['Code'] = 63
        row_list = []
        if each_row[Class_list_dict['isAComment']].value == 'N':
            row_class = each_row[Class_list_dict['Class name']].value
            row_sil_level = each_row[Class_list_dict['SIL Level']].value
            row_code = each_row[Class_list_dict['Code']].value

            for each in [row_class,row_code, row_sil_level]:
                row_list.append(each)
            class_list.append(row_list)

    for each_row in ws_2.rows:
        row_list = []
        if each_row[Instance_list_dict['isAComment']].value == 'N':
            row_id = ':R:A:' + judge_str_null(each_row[Instance_list_dict['Location']].value) + ':' + judge_str_null(
                each_row[Instance_list_dict['Sublocation']].value) + ':' + judge_str_null(each_row[Instance_list_dict['Equipment name']].value)

            row_class = each_row[Instance_list_dict['Class name']].value
            row_rtu_name = each_row[Instance_list_dict['RTU name']].value

            for each in [row_id, row_class, row_rtu_name]:
                row_list.append(each)
            ws_ev_list.append(row_list)

    for each_row in ws_4.rows:
        ev_class = each_row[EV_dict['Class name']].value
        ev_address = each_row[EV_dict['address']].value
        ev_deadband = each_row[EV_dict['deadband']].value
        ev_label = each_row[EV_dict['description']].value
        ev_length = each_row[EV_dict['length']].value
        ev_type = each_row[EV_dict['vetype']].value
        #ev_ptname = each_row[EV_dict['iopath']].value.replace(':dac', '')
        ev_ptname = each_row[EV_dict['iopath']].value
        ev_group = each_row[EV_dict['evgroup']].value
        ev_list = []
        if each_row[EV_dict['Class name']].value != 'Class name':
            for each in [ev_class, ev_ptname, ev_address, ev_deadband, ev_label, ev_length, ev_type, ev_group]:
                ev_list.append(each)
            ev_gp_list.append(ev_list)
    for each in ws_ev_list:

        class_name = each[1]
        nom_instance = 'link_ve'
        for each_ele in ev_gp_list:
            each_link = []
            each_ev = []
            deadband = each_ele[3]
            label = each_ele[4]
            length = each_ele[5]
            ev_tp = each_ele[6]
            print(label,length,ev_tp)
            pt_type = each_ele[1][0:3]
            pt_ele_type = type_dict[pt_type] + '_type_link_ve'
            address = each_ele[2]
            if each_ele[0] == class_name:
                link_id = each[0] + ':' + each_ele[1] + ':link_ve'
                for each_sil in class_list:
                    if each_sil[0] == each_ele[0] and each_sil[1] == str(each_ele[1]).split('-')[-1].split(':')[0]:
                        sil_type = sil_type_dict[each_sil[2]]
                        lk0 = ':R:A:POLE:' + each[2] + sil_type + ':' + each_ele[7] + ':' + ev_type_dict[pt_type] + str(
                            each[0][5:]).replace(':', '') + str(each_ele[1]).split('-')[-1].split(':')[0]

                print(link_id)
                each_link.append(nom_instance)
                each_link.append(link_id)
                each_link.append(pt_ele_type)
                each_link.append(lk0)
                each_ev.append(lk0)
                each_ev.append(str(lk0).split(':')[-1])
                each_ev.append('M-Conf ' + time.strftime("%d/%m/%y %H:%M"))
                each_ev.append(address)
                each_ev.append(deadband)
                each_ev.append(label)
                each_ev.append(length)
                each_ev.append(str(lk0).split(':')[-1])
                each_ev.append(ev_tp)
                each_ev.append('VE')
                link_list.append(each_link)
                ev_gen_list.append(each_ev)

    print(ws_ev_list)
    print(ev_gp_list)
    print(link_list)
    print(class_list)
    print(ev_gen_list)
    save_link_to_excel(link_list)
    save_ev_to_excel(ev_gen_list)


def gen_label(file_name):
    # file_name = 'LRT-QSTT-XX-ICD-CCS-GEN00-373543-B- SCADA-DPB-AppendixC_ATT1v0.37_HK_WP7.xlsx.xlsx'
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws_1 = wb['1-Class List']
    ws_2 = wb['2-Instance List']
    ws_3 = wb['3-Class Mapping']
    ws_4 = wb['4-External Variables']
    class_map_dict = {}
    for each_row in ws_3.rows:

        if each_row[0].value != 'Class name':
            class_map_dict[each_row[0].value] = each_row[1].value
    print(class_map_dict)
    total_row = ws_1.max_row
    print(total_row)
    # generate all class from ICD START
    for each_row in ws_1:
        if each_row[4].value is not None and each_row[4].value not in label_class_list:
            label_class_list.append(each_row[4].value)
    ### END

    for each_class in label_class_list:
        label_list = []
        hvid_list = []
        pt_list = []
        for each_row in ws_1:
            if each_row[4].value == each_class:
                if each_row[Class_list_dict['type']].value is not None and each_row[Class_list_dict['type']].value[1] == 'I':
                    try:
                        if each_row[Class_list_dict['Class name']].value != 'Class name':
                            str_pt_label = 'infopanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][9:] + '_' + each_row[Class_list_dict['code']].value[0].lower() + each_row[Class_list_dict['code']].value[1:] + ' : ' + each_row[Class_list_dict['CSS']].value + '\n'
                            pt_list.append(str_pt_label)
                        if each_row[Class_list_dict['bit_0']].value is not None:
                            str_bit1_label = 'infopanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                           9:] + '_' + each_row[Class_list_dict['code']].value[0].lower() + each_row[Class_list_dict['code']].value[1:] + '_0 : ' + \
                                           each_row[Class_list_dict['bit_0']].value + '\n'
                            pt_list.append(str_bit1_label)

                        if each_row[Class_list_dict['bit_1']].value is not None:
                            str_bit1_label = 'infopanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                           9:] + '_' + each_row[Class_list_dict['code']].value[0].lower() + each_row[Class_list_dict['code']].value[1:] + '_1 : ' + \
                                           each_row[Class_list_dict['bit_1']].value + '\n'
                            pt_list.append(str_bit1_label)

                        if each_row[Class_list_dict['bit_2']].value is not None:
                            str_bit1_label = 'infopanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                           9:] + '_' + each_row[Class_list_dict['code']].value[0].lower() + each_row[Class_list_dict['code']].value[1:] + '_2 : ' + \
                                           each_row[Class_list_dict['bit_2']].value + '\n'
                            pt_list.append(str_bit1_label)

                        if each_row[Class_list_dict['bit_3']].value is not None:
                            str_bit1_label = 'infopanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                           9:] + '_' + each_row[Class_list_dict['code']].value[0].lower() + each_row[Class_list_dict['code']].value[1:] + '_3 : ' + \
                                           each_row[Class_list_dict['bit_3']].value + '\n'
                            pt_list.append(str_bit1_label)
                    except KeyError:
                        print(each_row[4].value)
                elif each_row[Class_list_dict['type']].value is not None and each_row[Class_list_dict['type']].value[1] == 'O':
                    if each_row[Class_list_dict['SIL Level']].value == 'SIL0':
                        str_pt_label = 'cmdPanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                       9:] + '_' + each_row[Class_list_dict['code']].value + '_order : ' + \
                                       each_row[Class_list_dict['CSS']].value + '\n'
                        pt_list.append(str_pt_label)

                        if each_row[Class_list_dict['bit_0'] + 7].value != 'NONE' and each_row[Class_list_dict['bit_0'] + 7].value is not None:
                            str_bit1_label = 'cmdPanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                             9:] + '_' + each_row[Class_list_dict['code']].value + '_0 : ' + \
                                             each_row[Class_list_dict['bit_0'] + 7].value + '\n'
                            pt_list.append(str_bit1_label)

                        if each_row[Class_list_dict['bit_1'] + 7].value != 'NONE' and each_row[Class_list_dict['bit_1'] + 7].value is not None:
                            str_bit1_label = 'cmdPanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                             9:] + '_' + each_row[Class_list_dict['code']].value + '_1 : ' + \
                                             each_row[Class_list_dict['bit_1'] + 7].value + '\n'
                            pt_list.append(str_bit1_label)
                    else:
                        str_pt_label = 'cmdPanel_' + class_map_dict[each_row[Class_list_dict['Class name']].value][
                                                     9:] + '_' + each_row[Class_list_dict['code']].value[6:] + '_order : ' + \
                                       each_row[Class_list_dict['CSS']].value + '\n'
                        pt_list.append(str_pt_label)
        total_row1 = ws_2.max_row
        if total_row1 < total_row:
            total_row1 = total_row
        for row_num in range(3, total_row1):
            if ws_1['E' + str(row_num)].value is None and ws_1['E' + str(row_num + 1)].value == each_class:
                print(row_num, class_map_dict[ws_1['E' + str(row_num + 1)].value], ws_1['F' + str(row_num)].value)
                str_top1 = 'equipmentType_' + class_map_dict[ws_1['E' + str(row_num + 1)].value][9:] + '_std : ' + ws_1['F' + str(row_num)].value + '\n'
                str_top2 = class_map_dict[ws_1['E' + str(row_num + 1)].value][9:] + '_layer : ' + ws_1['F' + str(row_num)].value+ '\n'
                label_list.append('### ' + ws_1['E' + str(row_num + 1)].value + '\n')
                label_list.append('####################' + '\n')
                label_list.append(str_top1)
                label_list.append(str_top2)
                hvid_list.append('### ' + ws_1['E' + str(row_num + 1)].value + '\n')
                hvid_list.append('####################' + '\n')
            elif ws_1['E' + str(row_num)].value == each_class and ws_1['E' + str(row_num - 1)].value is None:
                for each_row in ws_2.rows:
                    if ws_1['E' + str(row_num)].value == each_row[8].value and each_row[22].value is not None:
                        str_lab = each_row[21].value + ' : ' + each_row[8].value + '\n'
                        hvid_lab = each_row[22].value + ' : ' + str(each_row[21].value).replace('&','') + '\n'
                        if str_lab not in label_list:
                            label_list.append(str_lab)
                            hvid_list.append(hvid_lab)
        all_label.append(label_list)
        all_label.append(pt_list)
        all_hvid.append(hvid_list)

    with open('label.txt', 'w') as f:
        for each_list in all_label:
            for each in each_list:
                f.write(each)

    with open('hvid.txt', 'w') as fhv:
        for each_list in all_hvid:
            for each in each_list:
                fhv.write(each)


def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = '{xlsx}.csv'.format(
                xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
                sheet=sheet.replace(' ', '_'))

            with open(csv_filename, 'w', newline='') as f_out:
                csv_file_writer = csv.writer(f_out)

                sheet_ranges = xlsx_file_reader[sheet]
                for row in sheet_ranges.rows:
                    row_container = []
                    for cell in row:
                        if cell.value is not None:
                            if type(cell.value) == str:
                                row_container.append(str(cell.value))
                            else:
                                row_container.append(str(cell.value))
                        else:
                            cell.value = ''
                            row_container.append(str(cell.value))
                    # print(row_container)
                    csv_file_writer.writerow(row_container)
        os.remove(filename)
    except Exception as e:
        print(e)


if __name__ == "__main__":
    print('Please enter your file name:\n')
    file_name = input()
    print('Please select: (Q for Quit)：')
    print('1. Add Equipment csv file.\n2. Add EV csv file.\n3. Generate HVID and labels\n')
    choice = input()
    while choice != 'Q':
        if choice == '1':
            equipment_csv(file_name)
            print('Generation Finish！')
            print('Please select,Q for Quit：')
            print('1. Add Equipment csv file.\n2. Add EV csv file.\n3. Generate HVID and labels\n')
            choice = input()
        elif choice == '2':
            ev(file_name)
            print('Generation Finish！')
            print('Please select,Q for Quit：')
            print('1. Add Equipment csv file.\n2. Add EV csv file.\n3. Generate HVID and labels\n')
            choice = input()
        elif choice == '3':
            gen_label(file_name)
            print('Generation Finish！')
            print('Please select,Q for Quit：')
            print('1. Add Equipment csv file.\n2. Add EV csv file.\n3. Generate HVID and labels\n')
            choice = input()
        else:
            print('Type error, please re-enter:')
            choice = input('Please select: (Q for Quit)：')

