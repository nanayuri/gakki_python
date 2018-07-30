#!/usr/bin/env python
# coding: utf-8


import os
import openpyxl
from openpyxl import load_workbook
from pypinyin import pinyin, lazy_pinyin
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import re
import time
import csv
from bidict import bidict
from bidict._exc import ValueDuplicationError
import shutil
import filecmp
import time

ev_name_dict = {
    'DI': 'dei',
    'AI': 'aei',
    'DO': 'deo',
    'AO': 'aeo',
    'SI': 'sei',
    'SO': 'seo'
}
ev_type_dict = {
    'DI': 'DEIV',
    'AI': 'AEIV',
    'DO': 'DEOV',
    'AO': 'AEOV',
    'SI': 'SEIV',
    'SO': 'SEOV'
}
fep_com_dict = {
    'HOM':  [208, 27],
    'PSCADA':  [160, 29],
    'BAS':  [123, 31],
    'FAS':  [32, 33],
    'PSD':  [33, 35],
    'AFC':  [34, 37],
    'DTS':  [35, 39],
    'ACS':  [36, 41],
    'PA':  [48, 43],
    'PIS':  [49, 45],
    'CCTV':  [50, 47],
    'ALM':  [51, 49],
    'FG01':  [96, 145],
    'FG02':  [97, 147],
    'UPS':  [112, 165],
    'ASD01':  [64, 105],
    'ASD02':  [65, 107],
    'ASD03':  [66, 109],
    'ASD04':  [67, 111],
    'ASD05':  [68, 113],
    'ASD06':  [69, 115],
    'ASD07':  [70, 117],
    'ASD08':  [71, 119],
    'ASD09':  [72, 121],
    'ASD10':  [73, 123],
    'ASD11':  [74, 125],
    'ASD12':  [75, 127],
    'ASD13':  [76, 129],
    'ASD14':  [77, 131],
    'ASD15':  [78, 133],
    'SIG_TRN':  [128, 185],
    'SIG_SWT':  [129, 187],
    'SIG_SIG':  [130, 189],
    'SIG_BLK':  [131, 191],
    'SIG_PLT':  [132, 193],
    'SIG_EMG':  [133, 195],
    'SIG_PID':  [134, 197],
    'IPL-Power':  [144, 205],
    'CLK':  [52, 51]
}

fep_com_sta_dict = {
    'HOM':  [208, 27],
    'PSCADA':  [160, 29],
    'BAS':  [123, 31],
    'FAS':  [32, 33],
    'PSD':  [33, 35],
    'AFC':  [34, 37],
    'DTS':  [35, 39],
    'ACS':  [36, 41],
    'PA':  [48, 43],
    'PIS':  [49, 45],
    'CCTV':  [50, 47],
    'ALM':  [51, 49],
    'FG01':  [96, 145],
    'FG02':  [97, 147],
    'UPS':  [112, 165],
    'ASD01':  [64, 105],
    'ASD02':  [65, 107],
    'ASD03':  [66, 109],
    'ASD04':  [67, 111],
    'ASD05':  [68, 113],
    'ASD06':  [69, 115],
    'ASD07':  [70, 117],
    'ASD08':  [71, 119],
    'ASD09':  [72, 121],
    'ASD10':  [73, 123],
    'ASD11':  [74, 125],
    'ASD12':  [75, 127],
    'ASD13':  [76, 129],
    'ASD14':  [77, 131],
    'ASD15':  [78, 133],
}

station_dict = {
    '高塘桥站': 'GTQ',
    '句章路站': 'JZL',
    '鄞州客运总站 ': 'YZC',
    '南部商务区站': 'SBD',
    '鄞州区政府站': 'YZG',
    '四明中路站': 'SMZ',
    '锦寓路站': 'JYL',
    '钱湖北路站': 'QHB',
    '仇毕站': 'QBZ',
    '儿童公园站': 'ETZ',
    '樱花公园站': 'YHZ',
    '体育馆站': 'TYG',
    '明楼站': 'MLZ',
    '中兴大桥南站': 'ZXN',
    '大通桥站': 'DTZ',
    '姜山站': 'JSZ',
    '狮山站': 'SSZ',
    '明辉路站': 'MHL',
    '朝阳站': 'CYZ',
    '方桥站': 'FQZ',
    '琎琳站': 'JLZ',
    '南渡站': 'NDZ',
    '大成东路站': 'DCZ',
    '金海路站': 'JHL',
    '首南车辆段': 'DEP',
    '奉化停车场': 'FHT',
    '控制中心': 'OCC',
    '樱花主变': 'YHB',
    '双桥主变': 'SQB',
    '蒋家主变电站': 'JJB'
}
list_name_dict = {}
list_name_dict['Station_Code'] = 'Station_Code\n车站名称'
list_name_dict['System'] = 'System\n系统'
list_name_dict['(Sub)System'] = '(Sub)System\n子系统'
list_name_dict['Eqpt_Code'] = 'Eqpt_Code\n设备类码'
list_name_dict['Eqpt_Desc'] = 'Eqpt_Description\n设备描述'
list_name_dict['Equipment_Location'] = 'Equipment_Location\n设备所在地'
list_name_dict['Eqpt_Identifier'] = 'Eqpt_Identifier\n设备标签'
list_name_dict['Attribute_Description'] = 'Attribute_Description\n属性描述'
list_name_dict['v0'] = 'v0_label(0)\n值0标签'
list_name_dict['v1'] = 'v1_label(1)\n值1标签'
list_name_dict['v2'] = 'v2_label(10)\n值2标签'
list_name_dict['v3'] = 'v3_label(11)\n值3标签'
list_name_dict['v4'] = 'v4_label(100)\n值4标签'
list_name_dict['v5'] = 'v5_label(101)\n值5标签'
list_name_dict['v6'] = 'v6_label(110)\n值6标签'
list_name_dict['v7'] = 'v7_label(111)\n值7标签'
list_name_dict['v0s'] = 'v0_Severity\n值0严重度'
list_name_dict['v1s'] = 'v1_Severity\n值1严重度'
list_name_dict['v2s'] = 'v2_Severity\n值2严重度'
list_name_dict['v3s'] = 'v3_Severity\n值3严重度'
list_name_dict['v4s'] = 'v4_Severity\n值4严重度'
list_name_dict['v5s'] = 'v5_Severity\n值5严重度'
list_name_dict['v6s'] = 'v6_Severity\n值6严重度'
list_name_dict['v7s'] = 'v7_Severity\n值7严重度'
list_name_dict['v0ic'] = 'v0_InitCond'
list_name_dict['v1ic'] = 'v1_InitCond'
list_name_dict['v2ic'] = 'v2_InitCond'
list_name_dict['v3ic'] = 'v3_InitCond'
list_name_dict['v0rc'] = 'v0_ReturnCond'
list_name_dict['v1rc'] = 'v1_ReturnCond'
list_name_dict['v2rc'] = 'v2_ReturnCond'
list_name_dict['v3rc'] = 'v3_ReturnCond'
list_name_dict['rcto'] = 'ReturnCond_TO\n反馈控制超时时间'
list_name_dict['HMI_Order'] = 'HMI_Order'
list_name_dict['CE_sig'] = 'Individual_identifier_of_(w)'
list_name_dict['DC_Data_Type'] = 'DC_Data_Type\nIO 类型'
list_name_dict['unit'] = 'unit\n单位'
list_name_dict['lower_limit'] = 'Value_Out_of_Range_LOW(i)\n报警下限'
list_name_dict['upper_limit'] = 'Value_Out_of_Range_HIGH(iv)\n报警上限'
list_name_dict['CFG_Equipment_Class'] = 'CFG_Equipment_Class'
list_name_dict['CFG_Element_Name'] = 'CFG_Element_Name'
list_name_dict['CFG_Data_Type'] = 'CFG_Data_Type'
list_name_dict['IV_Point_Name'] = 'IV_Point_Name'
list_name_dict['CFG_EQPT_ID'] = 'CFG_EQPT_ID'
list_name_dict['CFG_EQPT_ALIAS'] = 'CFG_EQPT_ALIAS'
list_name_dict['DB_Comment'] = 'DB_Comment'
list_name_dict['EV_Name'] = 'EV_Name'
list_name_dict['EV_Type'] = 'EV_Type'
list_name_dict['Jittor_Factor'] = 'Jittor_Factor'
list_name_dict['Transformation_Function'] = 'Transformation_Function'
list_name_dict['Deadband'] = '抖动量\nJittorfactor'
list_name_dict['Scaling'] = '模拟度量\nScaling factor'
list_name_dict['varInvalid'] = 'varInvalid1'
list_name_dict['ev_ID'] = 'CFG_EV_ID'
list_name_dict['ev_ADDRESS'] = 'SCADA_Address'
list_name_dict['swc_id'] = 'swc_id'
list_name_dict['table_id'] = 'table_id'
list_name_dict['start_byte'] = 'start_byte'
list_name_dict['start_bit'] = 'start_bit'
list_name_dict['FEP_addr_size'] = 'FEP_addr_size'

sys_dict = {}
sys_dict['PSCADA'] = 'POW'
sys_dict['BAS'] = 'BAS'
sys_dict['FAS'] = 'FAS'
sys_dict['ACS'] = 'ACS'
sys_dict['PIS'] = 'PIS'

# Define Common Parameters
cur_path = os.getcwd()
class_path = cur_path + '\\Class'
list_path = cur_path + '\\List'
alm_folder = cur_path + '\\mapping_file\\'
server_alm_po = '\\\\192.168.1.200\\d\\NBL3_ALMPO\\alm_cn.po'


def fill_list():
    out_path = cur_path + '\\Output_List'
    class_file_name = class_path + '\\' + os.listdir(class_path)[0]
    wb_class = openpyxl.load_workbook(class_file_name, data_only=True)
    ws_class = wb_class['CLASS']
    class_dict = {}
    i = 0
    for each in ws_class.columns:
        class_dict[each[1].value] = i
        i += 1
    max_row_num = ws_class.max_row
    while ws_class.cell(row=max_row_num, column=9).value is None:
        max_row_num -= 1
    class_list_ptname = {}
    class_list_eqptype = {}
    class_list_deadband = {}
    class_list_scaling = {}
    for row_num in range(3, max_row_num + 1):
        Eqpt_Code = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Eqpt_Code']] + 1).value
        Attribute_Description = ws_class.cell(row=row_num,
                                              column=class_dict[list_name_dict['Attribute_Description']] + 1).value
        DC_Data_Type = ws_class.cell(row=row_num, column=class_dict[list_name_dict['DC_Data_Type']] + 1).value
        point_name = ws_class.cell(row=row_num, column=class_dict[list_name_dict['IV_Point_Name']] + 1).value
        CFG_Element_Name = ws_class.cell(row=row_num, column=class_dict[list_name_dict['CFG_Element_Name']] + 1).value
        Deadband = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Deadband']] + 1).value
        Scaling = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Scaling']] + 1).value
        Eqp_num_str = Eqpt_Code + Attribute_Description + DC_Data_Type
        if Eqp_num_str not in class_list_ptname.keys():
            class_list_ptname[Eqp_num_str] = point_name
        if Eqp_num_str not in class_list_eqptype.keys():
            class_list_eqptype[Eqp_num_str] = CFG_Element_Name
        if Eqp_num_str not in class_list_deadband.keys():
            class_list_deadband[Eqp_num_str] = Deadband
        if Eqp_num_str not in class_list_scaling.keys():
            class_list_scaling[Eqp_num_str] = Scaling
    io_list_list = os.listdir(list_path)
    for each_file in io_list_list:
        file_name = list_path + '\\' + each_file
        # file_save_name = each_file[0:-5] + '_cfg.xlsx'
        file_save_path = out_path + '\\' + each_file
        wb = openpyxl.load_workbook(file_name, data_only=True)
        ws = wb['IO List']
        list_dict = {}
        i = 0
        for each in ws.columns:
            list_dict[each[1].value] = i
            i += 1
        list_eqp_num = {}
        max_row_num = ws.max_row
        while ws.cell(row=max_row_num, column=9).value is None:
            max_row_num -= 1
        for row_num in range(3, max_row_num + 1):
            Station_Code1 = ws.cell(row=row_num, column=list_dict[list_name_dict['Station_Code']] + 1).value
            Station_Code = station_dict[Station_Code1]
            System = ws.cell(row=row_num, column=list_dict[list_name_dict['System']] + 1).value
            Eqpt_Code = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Code']] + 1).value
            Eqpt_Desc = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Desc']] + 1).value
            Eqpt_Identifier = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Identifier']] + 1).value
            Attribute_Description = ws.cell(row=row_num,
                                            column=list_dict[list_name_dict['Attribute_Description']] + 1).value
            DC_Data_Type = ws.cell(row=row_num, column=list_dict[list_name_dict['DC_Data_Type']] + 1).value
            Deadband = ws.cell(row=row_num, column=class_dict[list_name_dict['Deadband']] + 1).value
            Scaling = ws.cell(row=row_num, column=class_dict[list_name_dict['Scaling']] + 1).value
            swc_id = "%(a)03d" % {'a' : int(ws.cell(row=row_num, column=class_dict[list_name_dict['swc_id']] + 1).value)}
            swc_id1 = ws.cell(row=row_num, column=class_dict[list_name_dict['swc_id']] + 1).value
            table_id = "%(a)03d" % {'a' : int(ws.cell(row=row_num, column=class_dict[list_name_dict['table_id']] + 1).value)}
            start_byte = "%(a)05d" % {'a': int(ws.cell(row=row_num, column=class_dict[list_name_dict['start_byte']] + 1).value)}
            start_bit = ws.cell(row=row_num, column=class_dict[list_name_dict['start_bit']] + 1).value
            FEP_addr_size = "%(a)06d" % {'a': int(ws.cell(row=row_num, column=class_dict[list_name_dict['FEP_addr_size']] + 1).value)}
            Eqp_num_str = Eqpt_Code + Attribute_Description + DC_Data_Type
            if Eqp_num_str not in list_eqp_num.keys():
                list_eqp_num[Eqp_num_str] = 1
            else:
                list_eqp_num[Eqp_num_str] += 1
            eqp_num = "%(a)04d" % {'a': list_eqp_num[Eqp_num_str]}
            # print(list_dict[list_name_dict['Eqpt_Code']])
            # print(row_num, ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Code']]).value)
            try:
                delete_cell = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Code']] + 1)
                if delete_cell.fill.fgColor.rgb == 'FFFF0000' and delete_cell.font.strike is True:
                    ws.cell(row=row_num, column=list_dict[list_name_dict['DB_Comment']] + 1).value = 'DELETE'
                else:
                    ws.cell(row=row_num, column=list_dict[list_name_dict['DB_Comment']] + 1).value = ''
                ws.cell(row=row_num,
                        column=list_dict[list_name_dict['CFG_Equipment_Class']] + 1).value = System + '_' + Eqpt_Code
                ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_Element_Name']] + 1).value = class_list_eqptype.setdefault(Eqp_num_str, '#N/A')
                ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_Data_Type']] + 1).value = DC_Data_Type
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'CFG_EQPT_ID']] + 1).value = ':R:A:' + Station_Code + ':' + System + ':' + Eqpt_Code + str(eqp_num)
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'CFG_EQPT_ALIAS']] + 1).value = '<alias>' + Station_Code + System + Eqpt_Code + str(eqp_num)
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'EV_Type']] + 1).value = ev_type_dict[DC_Data_Type[0:2]]
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'IV_Point_Name']] + 1).value = class_list_ptname.setdefault(Eqp_num_str, '#N/A')
                # print(Eqp_num_str)
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'ev_ADDRESS']] + 1).value = str(swc_id) + '/' + str(table_id) + '/' + str(start_byte) + '/' + str(start_bit) + '/' + str(FEP_addr_size)
                if ws.cell(row=row_num, column=list_dict[list_name_dict['IV_Point_Name']] + 1).value is None:
                    ws.cell(row=row_num, column=list_dict[list_name_dict[
                        'EV_Name']] + 1).value = ev_name_dict[
                                                     DC_Data_Type[0:2]] + Station_Code + System + Eqpt_Code + str(
                        eqp_num) + '-' + 'None' + str(row_num)
                else:
                    if ws.cell(row=row_num, column=list_dict[list_name_dict['IV_Point_Name']] + 1).value != '#N/A':
                        ws.cell(row=row_num, column=list_dict[list_name_dict[
                            'EV_Name']] + 1).value = ev_name_dict[
                                                         DC_Data_Type[0:2]] + Station_Code + System + Eqpt_Code + str(
                            eqp_num) + '-' + str(
                            ws.cell(row=row_num, column=list_dict[list_name_dict['IV_Point_Name']] + 1).value).split('-')[1]
                    else:
                        ws.cell(row=row_num, column=list_dict[list_name_dict[
                            'EV_Name']] + 1).value = '#N/A'
                ws.cell(row=row_num, column=list_dict[list_name_dict[
                    'ev_ID']] + 1).value = ':R:A:POLE_' + Station_Code + ':FEP:' + System + ':' + ws.cell(row=row_num,
                                                                                                          column=
                                                                                                          list_dict[
                                                                                                              list_name_dict[
                                                                                                                  'EV_Name']] + 1).value

                if DC_Data_Type[1] == 'I':
                    ws.cell(row=row_num, column=list_dict[list_name_dict[
                        'varInvalid']] + 1).value = 'dei' + Station_Code + 'SYSFEP' + fep_com_state(swc_id1) + '-COM'
                if Deadband is not None:
                    ws.cell(row=row_num, column=list_dict[list_name_dict[
                        'Jittor_Factor']] + 1).value = Deadband
                elif class_list_deadband.setdefault(Eqp_num_str, '#N/A') is not None and class_list_deadband.setdefault(Eqp_num_str, '#N/A') != '#N/A':
                    ws.cell(row=row_num, column=list_dict[list_name_dict[
                        'Jittor_Factor']] + 1).value = class_list_deadband.setdefault(Eqp_num_str, '#N/A')
                if Scaling is not None:
                    if type(Scaling) == int:
                        ws.cell(row=row_num, column=list_dict[list_name_dict[
                            'Transformation_Function']] + 1).value = 'tfLIN ' + str(Scaling) + '.0 0.0'
                    else:
                        ws.cell(row=row_num, column=list_dict[list_name_dict[
                            'Transformation_Function']] + 1).value = 'tfLIN ' + str(Scaling) + ' 0.0'
                elif class_list_scaling.setdefault(Eqp_num_str, '#N/A') is not None and class_list_scaling.setdefault(Eqp_num_str, '#N/A') != '#N/A':
                    ws.cell(row=row_num, column=list_dict[list_name_dict[
                        'Transformation_Function']] + 1).value = 'tfLIN ' + str(class_list_scaling.setdefault(Eqp_num_str, '#N/A')) + ' 0.0'
            except AttributeError:
                print(AttributeError)
        a1, a2 = add_rev_reg(wb)
        list_vis = 'V' + a1
        cur_date = a2
        str1 = str(re.match(r'(NB.*)(V.* )(\()(.*)(\))', each_file, re.M).group(2)).strip()
        str2 = str(re.match(r'(NB.*)(V.* )(\()(.*)(\))', each_file, re.M).group(4)).strip()
        wb.save(file_save_path.replace(str1, list_vis).replace(str2, cur_date))
        wb.close()


def gen_csv():
    list_out_path = cur_path + '\\Output_List'
    out_path = cur_path + '\\Generated_csv'
    alm_path = alm_folder + 'alm_cn.po'
    alm_save_path = alm_folder + 'new_alm.csv'
    io_list_list = os.listdir(list_out_path)
    alm_dict = read_alm_po(alm_path)[0]
    alm_content = read_alm_po(alm_path)[1]
    for each_file in io_list_list:
        file_name = list_out_path + '\\' + each_file
        equipment_csv_out = each_file.replace('.xlsx', '_equipment.xlsx')
        function_csv_out = each_file.replace('.xlsx', '_fcat.xlsx')
        ev_csv_out = each_file.replace('.xlsx', '_ev.xlsx')
        equipment_save_path = out_path + '\\' + equipment_csv_out
        ev_save_path = out_path + '\\' + ev_csv_out
        fct_save_path = out_path + '\\' + function_csv_out
        eqpt_list = []
        ev_list = []

        wb = openpyxl.load_workbook(file_name, data_only=True)
        ws = wb['IO List']
        list_dict = {}
        i = 0
        for each in ws.columns:
            list_dict[each[1].value] = i
            i += 1
        max_row_num = ws.max_row
        while ws.cell(row=max_row_num, column=9).value is None:
            max_row_num -= 1
        eq_list1 = []
        eq_list2 = []
        eq_list3 = []
        eq_list4 = []
        ev_list1 = []
        new_mapping_dict = bidict({})
        pow_sub_dict = {
            '35kV': ':R:A:Opm:Definition:Function:PSCADA-35',
            '400V': ':R:A:Opm:Definition:Function:PSCADA-400',
            '1500V': ':R:A:Opm:Definition:Function:PSCADA-1500'
        }
        pow_sub_list = []
        for row_num in range(3, max_row_num + 1):
            if ws.cell(row=row_num, column=list_dict[list_name_dict['DB_Comment']] + 1).value != 'DELETE':
                eq_row_list = []
                ev_row_list = []
                sub_sys_list = []
                Station_Code = ws.cell(row=row_num, column=list_dict[list_name_dict['Station_Code']] + 1).value
                sub_sys = ws.cell(row=row_num, column=list_dict[list_name_dict['(Sub)System']] + 1).value
                System = ws.cell(row=row_num, column=list_dict[list_name_dict['System']] + 1).value
                Eqpt_Code = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Code']] + 1).value
                Eqpt_Desc = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Desc']] + 1).value
                if Eqpt_Desc not in alm_dict:
                    tmp_res = word_length_chk(translation_to_pinyin(Eqpt_Desc))
                    tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                           Eqpt_Desc, tmp_res)
                    Eqpt_Desc = tmp_res
                else:
                    Eqpt_Desc = alm_dict[Eqpt_Desc]

                Eqpt_Identifier = ws.cell(row=row_num, column=list_dict[list_name_dict['Eqpt_Identifier']] + 1).value
                Attribute_Description = ws.cell(row=row_num,
                                                column=list_dict[list_name_dict['Attribute_Description']] + 1).value
                DC_Data_Type = ws.cell(row=row_num, column=list_dict[list_name_dict['DC_Data_Type']] + 1).value
                Deadband = ws.cell(row=row_num, column=list_dict[list_name_dict['Jittor_Factor']] + 1).value
                Scaling = ws.cell(row=row_num, column=list_dict[list_name_dict['Transformation_Function']] + 1).value
                FEP_addr_size = int(ws.cell(row=row_num, column=list_dict[list_name_dict['FEP_addr_size']] + 1).value)
                Eqp_num_str = Eqpt_Code + Attribute_Description + DC_Data_Type
                Eqpt_Location = ws.cell(row=row_num, column=list_dict[list_name_dict['Equipment_Location']] + 1).value
                if Eqpt_Location is not None:
                    if Eqpt_Location not in alm_dict:
                        tmp_res = word_length_chk(translation_to_pinyin(Eqpt_Location))
                        tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                               Eqpt_Location, tmp_res)
                        Eqpt_Location = tmp_res
                    else:
                        Eqpt_Location = alm_dict[Eqpt_Location]

                # Create Equipment Data to list eq_row_list
                Eq_ID = ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_EQPT_ID']] + 1).value
                Eq_fct_ID = Eq_ID + ':link_fct_cat'
                Eq_AGG = ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_Equipment_Class']] + 1).value
                Eq_Label = Eqpt_Identifier
                if Eq_Label is not None:
                    if Eq_Label not in alm_dict:
                        tmp_res = word_length_chk(translation_to_pinyin(Eq_Label))
                        tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                               Eq_Label, tmp_res)
                        Eq_Label = tmp_res
                    else:
                        Eq_Label = alm_dict[Eq_Label]

                Eq_Link = str(
                    ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_Element_Name']] + 1).value) + ':' + Eq_AGG
                Eq_Element = ws.cell(row=row_num, column=list_dict[list_name_dict['CFG_Element_Name']] + 1).value
                Eq_Lk0 = ''
                Eq_Lk1 = ''
                Eq_Short = Eqpt_Desc
                Eq_Loc = Eqpt_Location
                for each in [Eq_ID, Eq_AGG, Eq_Label, Eq_Link, Eq_Element, Eq_Lk0, Eq_Lk1, Eq_Short,
                             Eq_Loc]:
                    eq_row_list.append(each)
                if System == 'PSCADA':
                    if str(sub_sys) not in pow_sub_dict:
                        Eq_Sub = ':R:A:Opm:Definition:Function:POW'
                    else:
                        Eq_Sub = pow_sub_dict[str(sub_sys)]
                    for each in [Eq_fct_ID, Eq_Sub]:
                        sub_sys_list.append(each)
                    if sub_sys_list not in pow_sub_list:
                        pow_sub_list.append(sub_sys_list)
                else:
                    pow_sub_list = []

                if row_num == 3:
                    eq_list1.append(Eq_ID[0:8])
                    eq_list1.append('')
                    eq_list1.append(Eq_ID[5:8])
                    eq_list1.append('')
                    eq_list1.append('basicNode')
                    eq_list1.append('')
                    eq_list1.append('')
                    eq_list1.append('')
                    eq_list1.append('')
                    eq_list2.append(Eq_ID[0:8] + ':link_geo_cat')
                    eq_list2.append('')
                    eq_list2.append('')
                    eq_list2.append('')
                    eq_list2.append('')
                    eq_list2.append(':R:A:Opm:Definition:Location:' + Eq_ID[5:8])
                    eq_list2.append('')
                    eq_list2.append('')
                    eq_list2.append('')
                    eq_list3.append(':R:A:' + Eq_ID.split(':')[3] + ':' + Eq_ID.split(':')[4])
                    eq_list3.append('')
                    eq_list3.append(Eq_ID.split(':')[4])
                    eq_list3.append('')
                    eq_list3.append('basicNode')
                    eq_list3.append('')
                    eq_list3.append('')
                    eq_list3.append('')
                    eq_list3.append('')
                    eq_list4.append(':R:A:' + Eq_ID.split(':')[3] + ':' + Eq_ID.split(':')[4] + ':link_fct_cat')
                    eq_list4.append('')
                    eq_list4.append('')
                    eq_list4.append('')
                    eq_list4.append('')
                    if System == 'PSCADA':
                        eq_list4.append(':R:A:Opm:Definition:Function:POW')
                    else:
                        eq_list4.append(':R:A:Opm:Definition:Function:' + Eq_ID.split(':')[4])
                    eq_list4.append('')
                    eq_list4.append('')
                    eq_list4.append('')
                    eqpt_list.append(eq_list1)
                    eqpt_list.append(eq_list2)
                    eqpt_list.append(eq_list3)
                    eqpt_list.append(eq_list4)
                    eqpt_list.append(eq_row_list)
                else:
                    if eq_row_list not in eqpt_list:
                        eqpt_list.append(eq_row_list)

                # create EV data to list ev_row_list
                Ev_ID = ws.cell(row=row_num, column=list_dict[list_name_dict['ev_ID']] + 1).value
                Ev_Name = ws.cell(row=row_num, column=list_dict[list_name_dict['EV_Name']] + 1).value
                Ev_Address = ws.cell(row=row_num, column=list_dict[list_name_dict['ev_ADDRESS']] + 1).value
                Ev_Type = ws.cell(row=row_num, column=list_dict[list_name_dict['EV_Type']] + 1).value
                if FEP_addr_size <= 4 * 8:
                    Ev_Length = '4'
                elif FEP_addr_size <= 128 * 8:
                    Ev_Length = '128'
                elif FEP_addr_size <= 256 * 8:
                    Ev_Length = '256'
                elif FEP_addr_size <= 512 * 8:
                    Ev_Length = '512'
                Ev_Label = str(Eq_Label) + '-' + Eqpt_Desc
                if Deadband is not None:
                    Ev_deadband = Deadband
                else:
                    Ev_deadband = float('0.000')
                Ev_Element = 'VE'
                Ev_VarInvalid = ws.cell(row=row_num, column=list_dict[list_name_dict['varInvalid']] + 1).value
                Ev_Scaling = ws.cell(row=row_num, column=list_dict[list_name_dict['Transformation_Function']] + 1).value
                Ev_Sub = ''
                Ev_Ref = ''
                Ev_Link = ''
                for each in [Ev_ID, Ev_Name, Ev_Address, Ev_Type, Ev_Length, Ev_deadband, Ev_Label, Ev_Element,
                             Ev_VarInvalid, Ev_Scaling, Ev_Sub, Ev_Ref, Ev_Link]:
                    ev_row_list.append(each)
                if row_num == 3:
                    ev_list1.append(Ev_ID.replace(Ev_Name, '')[0:-1])
                    ev_list1.append(System)
                    ev_list1.append('')
                    ev_list1.append('')
                    ev_list1.append('')
                    ev_list1.append('')
                    ev_list1.append('')
                    ev_list1.append('GRP_VE')
                    ev_list1.append('')
                    ev_list1.append('')
                    ev_list1.append('0')
                    ev_list1.append('0')
                    ev_list1.append('')
                    ev_list.append(ev_list1)
                    ev_list.append(ev_row_list)
                else:
                    ev_list.append(ev_row_list)

        save_to_excel(eqpt_list, equipment_save_path)
        save_ev_to_excel(ev_list, ev_save_path)
        xlsx2csv(equipment_save_path)
        xlsx2csv(ev_save_path)
        save_new_po_csv(new_mapping_dict, alm_save_path)
        save_new_po(alm_content, new_mapping_dict, alm_path.replace('.po', '_new.po'))
        if len(pow_sub_list):
            save_function_to_excel(pow_sub_list, fct_save_path)
            xlsx2csv(fct_save_path)


def save_to_excel(data, file_name):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(
        ['ID', '__AGGREGATE__', 'label', 'link_object_is_from_object_type', 'ELEMENT_NAME', 'lk0', 'lk1', 'shortLabel',
         'EqtLocation'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def save_function_to_excel(data, file_name):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(
        ['ID', 'lk0'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def save_to_del_excel(data, file_name):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(
        ['ID', '__DELETE__'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def save_to_update_excel(data, file_name, type):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    if type == 'eq':
        ws_new.append(
            ['ID', 'label', 'EqtLocation'])
    elif type == 'db':
        ws_new.append(
            ['ID', 'functTrans', 'deadband'])
    elif type == 'addr':
        ws_new.append(
            ['ID', 'address', 'varInvalid1'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def save_ev_to_excel(data, file_name):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    ws_new.append(
        ['ID', 'name', 'address', 'type', 'length', 'deadband', 'label', 'ELEMENT_NAME', 'varInvalid1', 'functTrans',
         'SubscriptionPeriod', 'RefreshPeriod', 'link_object_is_from_object_type'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def save_class_excel(data, file_name, csv_type):
    wb_new = openpyxl.Workbook()
    wb_new.guess_types = True
    ws_new = wb_new.active
    if csv_type == 'eq':
        ws_new.append(
            ['ID', 'ELEMENT_NAME', 'shortLabel', 'type'])
    elif csv_type == 'pt':
        ws_new.append(
            ['ID', 'ELEMENT_NAME', 'label', 'valueLength', 'unit', '__AGGREGATE__', 'initCond1', 'returnCond1', 'returnCondTO', 'Archive', 'hmiOrder', 'computedMessage', 'type'])
    elif csv_type == 'xac':
        ws_new.append(
            ['ID', 'ELEMENT_NAME', 'nature'])
    elif csv_type == 'xal':
        ws_new.append(
            ['ID', 'ELEMENT_NAME', 'ackAutomaton', 'valueAutomaton', 'theme'])
    elif csv_type == 'xfo':
        ws_new.append(
            ['ID', 'ELEMENT_NAME'])
    elif csv_type == 'dov':
        ws_new.append(
            ['ID', 'ELEMENT_NAME', 'ctlvalue', 'initCond1', 'returnCond1', 'returnCondTO'])
    elif csv_type == 'dioVT':
        ws_new.append(
            ['ID', 'v0', 'v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'])
    elif csv_type == 'dalVT':
        ws_new.append(
            ['ID', 'v0', 'v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7'])
    elif csv_type == 'dalVT10':
        ws_new.append(
            ['ID', 'v0', 'v1', 'v2', 'v3', 'v4', 'v5', 'v6', 'v7', 'v8', 'v9'])
    elif csv_type == 'aalVT':
        ws_new.append(
            ['ID', 'v0', 'v1', 'v2', 'v3', 'v4'])
    elif csv_type == 'aalVL':
        ws_new.append(
            ['ID', 'v0', 'v1', 'v2', 'v3', 'v4'])
    for each in data:
        ws_new.append(each)
    wb_new.save(file_name)
    wb_new.close()


def translation_to_pinyin(str1):
    if str1 is not None:
        if type(str1) != type('abc'):
            str1 = str(str1)
        str1 = str1.strip().replace('℃', 'degree').replace('³', '3').replace('%', 'percent')
        str1 = re.sub('\W*', '', str1)
    result_list = lazy_pinyin(str1)
    result = '&DB_'
    for each in result_list:
        if each != ' ':
            # print(re.sub("\W*", "", each[4:].replace('_', '')))
            each = (each[0:4] + re.sub("\W*", "", each[4:].replace('_', '').replace('℃', 'C').replace('Ω', ''))).upper()
            result += each
    return result


def word_length_chk(str1):
    result = ''
    if len(str1) >= 58:
        result = str1[:58]
    else:
        result = str1
    return result


def word_length10_chk(str1):
    result = ''
    if len(str1) > 8:
        result = str1[:8]
    else:
        result = str1
    return result


def word_length5_chk(str1):
    result = ''
    str1 = str1.replace('&DB_', '')
    if len(str1) > 5:
        result = str1[:5]
    else:
        result = str1
    return result


def word_length_chk32(str1):
    result = ''
    if len(str1) >= 26:
        result = str1[:26]
    else:
        result = str1
    return result


def read_alm_po(file_path):
    with open(file_path, 'r', encoding='UTF-8') as f1:
        content = f1.readlines()
    chi_list = []
    eng_list = []
    alm_dict = bidict()
    bi_alm_dict = {}

    for each_line in content:
        chi_word = ''
        eng_word = ''
        if 'msgstr' in each_line:
            matchObj = re.match(r'(msgstr )"(.*)"', each_line, re.M)
            chi_word = matchObj.group(2)

        if chi_word != '\"\"' and chi_word != '\"\"\n' and chi_word != '':
            chi_list.append(chi_word)
        if 'msgid' in each_line:
            matchObj = re.match(r'(msgid )"(.*)"', each_line, re.M)
            eng_word = matchObj.group(2)
        if eng_word != '\"\"' and eng_word != '\"\"\n' and eng_word != '':
            eng_list.append(eng_word)
    if len(chi_list) != len(eng_list):
        print('mapping file error')
    else:
        # print(eng_list, chi_list)
        # print(len(chi_word))
        for i in range(0, len(chi_list)):
            try:
                alm_dict[chi_list[i].replace('\n', '').replace('"', '')] = eng_list[i].replace('\n', '').replace('"', '')
            except ValueDuplicationError:
                alm_dict[chi_list[i].replace('\n', '').replace('"', '')] = eng_list[i].replace('\n', '').replace('"',
                                                                                                                 '') + '_1'
    return alm_dict, content


def save_new_po_csv(dict1, file_path):
    with open(file_path, 'w') as f1:
        for each_line in dict1.items():
            try:
                f1.write(str(each_line[0]) + ',' + str(each_line[1]) + '\n')
            except UnicodeEncodeError:
                print(each_line[0] + ',' + each_line[1])


def save_new_po(content, dict1, file_path):
    with open(file_path, 'w', encoding='utf-8') as f1:
        f1.writelines(content)
        f1.write('\n')
        f1.write('\n')
        for each_line in dict1.items():
            f1.write('msgid "' + str(each_line[1]) + '"' + '\n')
            f1.write('msgstr "' + str(each_line[0]) + '"' + '\n')
            f1.write('\n')


def gen_class_csv():
    out_path = cur_path + '\\Class_csv'
    alm_path = alm_folder + 'alm_cn.po'
    alm_save_path = alm_folder + 'new_class_alm.csv'
    Internal_without_pt_list = ['dciPOW-MAINT', 'dciPOW-INHIBIT', 'dciPOW-EARTH', 'dciBAS-HOSTAT']
    Internal_without_ptdo_list = ['dioPOW-MAINT', 'dioPOW-INHIBIT', 'dioPOW-EARTH', 'dioBAS-HOSTAT']
    csv_1_eq_path = out_path + '\\' + '1-Equipment.xlsx'
    csv_2_pt_path = out_path + '\\' + '2-PointInstance.xlsx'
    csv_3_xac_path = out_path + '\\' + '3-XAC.xlsx'
    csv_3_xal_path = out_path + '\\' + '3-XAL.xlsx'
    csv_3_xfo_path = out_path + '\\' + '3-XFO.xlsx'
    csv_3_dov_path = out_path + '\\' + '3-dov.xlsx'
    csv_4_dio_path = out_path + '\\' + '4-DioValueTable.xlsx'
    csv_4_dal_path = out_path + '\\' + '4-DalValueTable.xlsx'
    csv_4_dal10_path = out_path + '\\' + '4-Dal10ValueTable.xlsx'
    csv_4_aal_path = out_path + '\\' + '4-AalValueTable.xlsx'
    csv_4_aalvl_path = out_path + '\\' + '4-AalValueLimits.xlsx'
    class_file_name = class_path + '\\' + os.listdir(class_path)[0]
    aci_ce = 'SCS_FTOA([.value],"2",[.unit])'
    ao_ce = 'Int2String(Byte2Int([.value],0,2))'
    do_ce = 'Int2String(Byte2Int([.value],0,2))'
    wb_class = openpyxl.load_workbook(class_file_name, data_only=True)
    ws_class = wb_class['CLASS']
    class_dict = {}
    alm_dict = read_alm_po(alm_path)[0]
    alm_content = read_alm_po(alm_path)[1]
    i = 0
    for each in ws_class.columns:
        class_dict[each[1].value] = i
        i += 1
    max_row_num = ws_class.max_row
    while ws_class.cell(row=max_row_num, column=9).value is None:
        max_row_num -= 1
    class_list_ptname = {}
    class_list_eqptype = {}
    class_list_deadband = {}
    class_list_scaling = {}
    new_mapping_dict = bidict({})
    class_eq_list = []
    class_pt_list = []
    class_xac_list = []
    class_xal_list = []
    class_xfo_list = []
    class_dov_list = []
    class_dio_list = []
    class_dal_list = []
    class_dal10_list = []
    class_aal_list = []
    class_aal_vl_list = []
    pt_type_dict = {}
    pt_type_dict['AI'] = 'aci_cb_type'
    pt_type_dict['DI1/SOE'] = 'dci_soe_type'
    pt_type_dict['DI1'] = 'dci_cb_type'
    pt_type_dict['DI2'] = 'dci_cb_type'
    pt_type_dict['DI'] = 'dci_cb_type'
    pt_type_dict['DO2'] = 'dio_type'
    pt_type_dict['DO'] = 'dio_type'
    pt_type_dict['DI_INT'] = 'dci_cb_type'
    pt_type_dict['DO_INT'] = 'dio_type'
    pt_type_dict['AO'] = 'aio_type'
    pt_type_dict['SI'] = 'sci_cb_type'
    pt_type_dict['SO'] = 'sio_type'
    pt_type_dict['AI2'] = 'aci_cb_type'
    pt_type_dict['DO1'] = 'dio_type'

    xac_type_dict = {}
    xac_type_dict['AI'] = 'aac_type'
    xac_type_dict['DI1/SOE'] = 'dac_soe_type'
    xac_type_dict['DI1'] = 'dac_type'
    xac_type_dict['DI2'] = 'dac_type'
    xac_type_dict['DI'] = 'dac_type'
    xac_type_dict['DI_INT'] = 'dac_type'
    xac_type_dict['SI'] = 'sac_type'
    xac_type_dict['AI2'] = 'aac_type'
    xac_appendix_dict = {}
    xac_appendix_dict['dci'] = 'dac'
    xac_appendix_dict['aci'] = 'aac'
    xac_appendix_dict['sci'] = 'sac'

    xal_type_dict = {}
    xal_type_dict['AI'] = 'aal_type'
    xal_type_dict['AI2'] = 'aal_type'
    xal_type_dict['DI1/SOE'] = 'dal_type'
    xal_type_dict['DI1'] = 'dal_type'
    xal_type_dict['DI2'] = 'dal_type'
    xal_type_dict['DI'] = 'dal_type'
    xal_type_dict['DI_INT'] = 'dal_10_type'
    xal_appendix_dict = {}
    xal_appendix_dict['dci'] = 'dal'
    xal_appendix_dict['aci'] = 'aal'

    for row_num in range(3, max_row_num + 1):
        System = ws_class.cell(row=row_num, column=class_dict[list_name_dict['System']] + 1).value
        Eqpt_Code = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Eqpt_Code']] + 1).value
        CE_SIG = str(ws_class.cell(row=row_num, column=class_dict[list_name_dict['CE_sig']] + 1).value)


        DC_Data_Type = ws_class.cell(row=row_num, column=class_dict[list_name_dict['DC_Data_Type']] + 1).value
        v0 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v0']] + 1).value
        v1 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v1']] + 1).value
        v2 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v2']] + 1).value
        v3 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v3']] + 1).value
        v4 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v4']] + 1).value
        v5 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v5']] + 1).value
        v6 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v6']] + 1).value
        v7 = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v7']] + 1).value

        v0s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v0s']] + 1).value
        v1s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v1s']] + 1).value
        v2s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v2s']] + 1).value
        v3s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v3s']] + 1).value
        v4s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v4s']] + 1).value
        v5s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v5s']] + 1).value
        v6s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v6s']] + 1).value
        v7s = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v7s']] + 1).value

        v0ic = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v0ic']] + 1).value
        v1ic = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v1ic']] + 1).value
        v2ic = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v2ic']] + 1).value
        v3ic = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v3ic']] + 1).value
        v0rc = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v0rc']] + 1).value
        v1rc = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v1rc']] + 1).value
        v2rc = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v2rc']] + 1).value
        v3rc = ws_class.cell(row=row_num, column=class_dict[list_name_dict['v3rc']] + 1).value
        rcto = ws_class.cell(row=row_num, column=class_dict[list_name_dict['rcto']] + 1).value

        point_name = ws_class.cell(row=row_num, column=class_dict[list_name_dict['IV_Point_Name']] + 1).value
        CFG_Element_Name = ws_class.cell(row=row_num, column=class_dict[list_name_dict['CFG_Element_Name']] + 1).value
        CFG_Equipment_Class = ws_class.cell(row=row_num, column=class_dict[list_name_dict['CFG_Equipment_Class']] + 1).value
        HMI_Order = ws_class.cell(row=row_num, column=class_dict[list_name_dict['HMI_Order']] + 1).value
        Unit = ws_class.cell(row=row_num, column=class_dict[list_name_dict['unit']] + 1).value
        if Unit is not None:
            if Unit not in alm_dict:
                tmp_res = '&DB'
                tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                       Unit, tmp_res)

                Unit = tmp_res
            else:
                Unit = alm_dict[Unit]

        Deadband = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Deadband']] + 1).value
        Scaling = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Scaling']] + 1).value
        Eq_Id = ':R:A:' + System + ':' + Eqpt_Code
        Pt_Id = Eq_Id + ':' + str(point_name)



        # 4-DioValueTable
        dio_table_dict = {0: '', 1: '', 2: '', 3: ''}
        if DC_Data_Type[0:2] == 'DO':
            dio_table_dict[0] = v0
            dio_table_dict[1] = v1
            dio_table_dict[2] = v2
            dio_table_dict[3] = v3
            vt_name = ':V_valueTable_'
            for each in ['dovname', 'label', 'value']:
                dio_each_list = []
                if each == 'dovname':
                    dio_ID = Pt_Id + vt_name + each
                    dio_each_list.append(dio_ID)
                    if DC_Data_Type[-3:] == 'INT' and point_name not in Internal_without_ptdo_list:
                        for each_value in dio_table_dict.items():
                            if each_value[1] is not None:
                                dov_name = 'dovPOW-' + str(each_value[1])
                                dio_each_list.append(dov_name)
                        class_dio_list.append(dio_each_list)
                    else:
                        for each_value in dio_table_dict.items():
                            if each_value[1] is not None:
                                dov_name = 'dov' + sys_dict[System] + '-CTLVAL' + str(each_value[0])
                                dio_each_list.append(dov_name)
                        class_dio_list.append(dio_each_list)
                elif each == 'label':
                    dio_ID = Pt_Id + vt_name + each
                    dio_each_list.append(dio_ID)
                    for each_label in [v0, v1, v2, v3]:
                        if each_label is not None:
                            if each_label not in alm_dict:
                                tmp_res = word_length10_chk(translation_to_pinyin(each_label))
                                tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                                       each_label, tmp_res)
                                each_label = tmp_res

                            else:
                                each_label = alm_dict[each_label]
                            dio_each_list.append(each_label)
                    class_dio_list.append(dio_each_list)
                elif each == 'value':
                    dio_ID = Pt_Id + vt_name + each
                    dio_each_list.append(dio_ID)
                    for each_value in dio_table_dict.items():
                        if each_value[1] is not None:
                            dio_each_list.append(each_value[0])
                    class_dio_list.append(dio_each_list)
        # 4-DalValueTable
        dal_table_dict = {0: '', 1: '', 2: '', 3: '', 4: '', 5: '', 6: '', 7: ''}
        dal_table_severity_dict = {0: '', 1: '', 2: '', 3: '', 4: '', 5: '', 6: '', 7: ''}
        dal10_table_dict = {0: '', 1: '', 2: '', 3: ''}
        if DC_Data_Type[0:2] == 'DI':
            dal_vt_name = ':dal:V_valueTable_'
            if DC_Data_Type[-3:] == 'INT' and point_name not in Internal_without_pt_list:
                dal_table_dict = {0: '&DB_DESELECT_SUCCESS', 1: '&DB_DESELECT_FAIL', 2: '&DB_DESELECT_UNKNOWN',
                                  3: '&DB_EXECUTE_SUCCESS', 4: '&DB_EXECUTE_FAIL', 5: '&DB_EXECUTE_UNKNOWN',
                                  6: '&DB_SELECT_SUCCESS', 7: '&DB_SELECT_FAIL', 8: '&DB_SELECT_UNKNOWN', 9: '&DB_EQUIPMENT_INIT'}
                dal_table_severity_dict = {0: '0', 1: '0', 2: '0',
                                           3: '0', 4: '0', 5: '0',
                                           6: '0', 7: '0', 8: '0', 9: '0'}
                dal_table_value_dict = {0: '1', 1: '2', 2: '3',
                                           3: '4', 4: '8', 5: '12',
                                           6: '16', 7: '32', 8: '48', 9: '0'}
                for each in ['format', 'label', 'severity', 'state', 'value']:
                    dal_each_list = []
                    if each == 'format':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_dict.items():
                            if each_cat[1] is not None and each_cat[1] != '':
                                if 'SOE' in DC_Data_Type:
                                    dal_each_list.append(2)
                                else:
                                    dal_each_list.append(0)

                    if each == 'label':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_dict.items():
                            dal_label = each_cat[1]
                            dal_each_list.append(dal_label)
                    if each == 'severity':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_severity_dict.items():
                            dal_sev = each_cat[1]
                            if dal_sev is not None and each_cat[1] != '':
                                dal_each_list.append(dal_sev)

                    if each == 'state':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_severity_dict.items():
                            dal_sev = each_cat[1]
                            if dal_sev is not None and each_cat[1] != '':
                                if int(dal_sev) != 0:
                                    dal_each_list.append('A')
                                else:
                                    dal_each_list.append('N')

                    if each == 'value':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_value_dict.items():
                            dal_label = each_cat[1]
                            if dal_label is not None and each_cat[1] != '':
                                dal_each_list.append(each_cat[1])
                    class_dal10_list.append(dal_each_list)
            else:
                dal_table_dict[0] = v0
                dal_table_dict[1] = v1
                dal_table_dict[2] = v2
                dal_table_dict[3] = str(v3).replace(' ', '').replace('None', '')
                dal_table_dict[4] = v4
                dal_table_dict[5] = v5
                dal_table_dict[6] = v6
                dal_table_dict[7] = v7
                dal_table_severity_dict[0] = v0s
                dal_table_severity_dict[1] = v1s
                dal_table_severity_dict[2] = v2s
                dal_table_severity_dict[3] = str(v3s).replace(' ', '').replace('None', '')
                dal_table_severity_dict[4] = v4s
                dal_table_severity_dict[5] = v5s
                dal_table_severity_dict[6] = v6s
                dal_table_severity_dict[7] = v7s

                for each in ['format', 'label', 'severity', 'state', 'value']:
                    dal_each_list = []
                    if each == 'format':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_dict.items():
                            if each_cat[1] is not None and each_cat[1] != '':
                                if 'SOE' in DC_Data_Type:
                                    dal_each_list.append(2)
                                else:
                                    dal_each_list.append(0)

                    if each == 'label':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_dict.items():
                            dal_label = each_cat[1]
                            if dal_label is not None and each_cat[1] != '':
                                # print(row_num)
                                if dal_label not in alm_dict:
                                    tmp_res = word_length10_chk(translation_to_pinyin(dal_label))
                                    tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict, dal_label, tmp_res)
                                    dal_label = tmp_res

                                else:
                                    dal_label = alm_dict[dal_label]
                                dal_each_list.append(dal_label)
                    if each == 'severity':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_severity_dict.items():
                            dal_sev = each_cat[1]
                            if dal_sev is not None and each_cat[1] != '':
                                dal_each_list.append(dal_sev)

                    if each == 'state':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_severity_dict.items():
                            dal_sev = each_cat[1]
                            if dal_sev is not None and each_cat[1] != '':
                                if int(dal_sev) != 0:
                                    dal_each_list.append('A')
                                else:
                                    dal_each_list.append('N')

                    if each == 'value':
                        dal_ID = Pt_Id + dal_vt_name + each
                        dal_each_list.append(dal_ID)
                        for each_cat in dal_table_dict.items():
                            dal_label = each_cat[1]
                            if dal_label is not None and each_cat[1] != '':
                                dal_each_list.append(each_cat[0])

                    class_dal_list.append(dal_each_list)

        Attribute_Description = ws_class.cell(row=row_num,
                                              column=class_dict[list_name_dict['Attribute_Description']] + 1).value
        Attribute_Description = str(Attribute_Description).strip()
        if Attribute_Description not in alm_dict:
            tmp_res = word_length_chk(translation_to_pinyin(Attribute_Description))
            tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                   Attribute_Description, tmp_res)
            Attribute_Description = tmp_res
        else:
            Attribute_Description = alm_dict[Attribute_Description]

        Eqpt_Desc = ws_class.cell(row=row_num, column=class_dict[list_name_dict['Eqpt_Desc']] + 1).value
        if Eqpt_Desc not in alm_dict:
            tmp_res = word_length_chk32(translation_to_pinyin(Eqpt_Desc))
            tmp_res, new_mapping_dict, alm_dict = insert_dup_value(new_mapping_dict, alm_dict,
                                                                   Eqpt_Desc, tmp_res)
            Eqpt_Desc = tmp_res
        else:
            Eqpt_Desc = alm_dict[Eqpt_Desc]
        # 1-Equipments.csv
        eq_list = []
        if row_num == 3:
            eq_list.append(':R:A:' + System)
            eq_list.append('basicNode')
            eq_list.append('')
            eq_list.append('')
            class_eq_list.append(eq_list)
            eq_list = []
        eq_list.append(Eq_Id)
        eq_list.append(CFG_Element_Name)
        eq_list.append(Eqpt_Desc)
        # eq_list.append(CFG_Equipment_Class)
        if eq_list not in class_eq_list:
            class_eq_list.append(eq_list)
        # 2-PointInstance.csv
        pt_list = []
        pt_list.append(Pt_Id)
        pt_list.append(pt_type_dict[DC_Data_Type])
        pt_list.append(Attribute_Description)
        pt_list.append('')
        pt_list.append(Unit)
        pt_list.append('')
        pt_list.append('')
        pt_list.append('')
        pt_list.append('')
        pt_list.append('')
        pt_list.append(HMI_Order)
        if point_name[0:3] == 'aci':
            pt_list.append(aci_ce)
        elif point_name[0:3] == 'aio':
            pt_list.append(ao_ce)
        elif point_name[0:3] == 'dio' or point_name[0:3] == 'sio':
            pt_list.append(do_ce)
        else:
            pt_list.append('')
        if point_name[0:3] == 'dci':
            if CE_SIG == '1':
                pt_list.append('1')
            elif CE_SIG == '2':
                pt_list.append('2')
            else:
                pt_list.append('__ERASE__')
        else:
            pt_list.append('')
        class_pt_list.append(pt_list)
        # 3-XAC.csv
        xac_list = []
        if DC_Data_Type[1] == 'I':
            xac_list.append(Pt_Id + ':' + str(xac_appendix_dict[point_name[0:3]]))
            xac_list.append(xac_type_dict[DC_Data_Type])
            if point_name[0] == 'd':
                xac_list.append('1')
            else:
                xac_list.append('')
            if xac_list not in class_xac_list:
                class_xac_list.append(xac_list)
        # 3-XAL.csv
        xal_list = []
        if DC_Data_Type[1] == 'I':
            if DC_Data_Type[0] != 'S':
                xal_list.append(Pt_Id + ':' + str(xal_appendix_dict[point_name[0:3]]))
                if DC_Data_Type[-3:] == 'INT':
                    if point_name not in Internal_without_pt_list:
                        xal_list.append(xal_type_dict[DC_Data_Type])
                    else:
                        xal_list.append('dal_type')
                else:
                    xal_list.append(xal_type_dict[DC_Data_Type])
                xal_list.append('Ack')
                if point_name[0] == 'd':
                    xal_list.append('DIBasic2StateLog')
                else:
                    xal_list.append('AIBasic2StateHisterisis')
                xal_list.append('0')
                if xal_list not in class_xal_list:
                    class_xal_list.append(xal_list)
        # 3-XFO.csv
        xfo_list = []
        if DC_Data_Type[1] == 'I':
            if DC_Data_Type[0] == 'D':
                xfo_list.append(Pt_Id + ':' + 'dfo')
                xfo_list.append('dfo_type')
            elif DC_Data_Type[0] == 'A':
                xfo_list.append(Pt_Id + ':' + 'afo')
                xfo_list.append('afo_type')
            else:
                xfo_list.append(Pt_Id + ':' + 'sfo')
                xfo_list.append('sfo_type')
            if xfo_list not in class_xfo_list:
                class_xfo_list.append(xfo_list)
        # 3-dov.csv
        dov_table_dict = {0:'', 1:'', 2:'', 3:''}
        num_list = []
        dov_name_list = []
        if DC_Data_Type[0:2] == 'DO':
            dov_table_dict[0] = v0
            dov_table_dict[1] = v1
            dov_table_dict[2] = v2
            # dov_table_dict[3] = str(v3).strip()
            dov_table_dict[3] = v3
            if DC_Data_Type[-3:] != 'INT' or point_name in Internal_without_ptdo_list:
                for i in range(0, 4):
                    if dov_table_dict[i] is not None:
                        num_list.append(i)
                for each in num_list:
                    dov_list = []
                    dov_name_last = 'dov' + sys_dict[System] + '-CTLVAL' + str(each)
                    dov_ID = Pt_Id + ':' + dov_name_last
                    dov_list.append(dov_ID)
                    dov_list.append('dov_type')
                    if System == 'PSCADA':
                        dov_list.append(8192 + int(each))
                    else:
                        dov_list.append(int(each))
                    if each == 0:
                        dov_list.append(v0ic)
                        dov_list.append(v0rc)
                        if v0rc != '':
                            dov_list.append(rcto)
                    elif each == 1:
                        dov_list.append(v1ic)
                        dov_list.append(v1rc)
                        if v1rc != '':
                            dov_list.append(rcto)
                    elif each == 2:
                        dov_list.append(v2ic)
                        dov_list.append(v2rc)
                        if v2rc != '':
                            dov_list.append(rcto)
                    elif each == 3:
                        dov_list.append(v3ic)
                        dov_list.append(v3rc)
                        if v3rc != '':
                            dov_list.append(rcto)
                    else:
                        dov_list.append('')
                        dov_list.append('')
                        dov_list.append('')
                    if dov_list not in class_dov_list:
                        class_dov_list.append(dov_list)
            else:
                num = 0

                for i in range(0, 4):
                    if dov_table_dict[i] is not None:
                        num += 1
                if num == 2:
                    for each in dov_table_dict.items():
                        dov_list = []
                        if each[1] is not None:
                            dov_name_last = 'dovPOW-' + str(each[1])
                            dov_ID = Pt_Id + ':' + dov_name_last
                            dov_list.append(dov_ID)
                            dov_list.append('dov_type')
                            if each[0] == 0:
                                dov_list.append(4096 + 2)
                                dov_list.append('')
                                dov_list.append('')
                                dov_list.append('')
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
                            else:
                                dov_list.append(16384 + 2)
                                dov_list.append('')
                                dov_list.append('')
                                dov_list.append('')
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
                elif num == 4:
                    for each in dov_table_dict.items():
                        dov_list = []
                        if each[1] is not None:
                            dov_name_last = 'dovPOW-' + str(each[1])
                            dov_ID = Pt_Id + ':' + dov_name_last
                            dov_list.append(dov_ID)
                            dov_list.append('dov_type')
                            if each[0] == 0:
                                dov_list.append(4096 + 1)
                                dov_list.append(v0ic)
                                dov_list.append(v0rc)
                                if v0rc != '':
                                    dov_list.append(rcto)
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
                            elif each[0] == 1:
                                dov_list.append(4096 + 2)
                                dov_list.append(v1ic)
                                dov_list.append(v1rc)
                                if v1rc != '':
                                    dov_list.append(rcto)
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
                            elif each[0] == 2:
                                dov_list.append(16384 + 1)
                                dov_list.append(v2ic)
                                dov_list.append(v2rc)
                                if v2rc != '':
                                    dov_list.append(rcto)
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
                            elif each[0] == 3:
                                dov_list.append(16384 + 2)
                                dov_list.append(v3ic)
                                dov_list.append(v3rc)
                                if v3rc != '':
                                    dov_list.append(rcto)
                                if dov_list not in class_dov_list:
                                    class_dov_list.append(dov_list)
         # 4-AalValueTable
        aal_table_dict = {0: '&DB_OUTOFRANGELOW', 1: '&DB_LOW', 2: '&DB_NORMAL', 3: '&DB_HIGH',
                          4: '&DB_OUTOFRANGEHIGH'}
        aal_table_severity_dict = {0: '4', 1: '4', 2: '0', 3: '4', 4: '4'}

        if DC_Data_Type[0:2] == 'AI':
            aal_vt_name = ':aal:V_valueTable_'
            for each in ['format', 'index', 'label', 'severity', 'state']:
                aal_each_list = []
                if each == 'format':
                    aal_ID = Pt_Id + aal_vt_name + each
                    aal_each_list.append(aal_ID)
                    for each_cat in aal_table_dict.items():
                        if each_cat[1] is not None and each_cat[1] != '':
                            aal_each_list.append(1)
                if each == 'index':
                    aal_ID = Pt_Id + aal_vt_name + each
                    aal_each_list.append(aal_ID)
                    for each_cat in aal_table_dict.items():
                        aal_label = each_cat[1]
                        if aal_label is not None and each_cat[1] != '':
                            aal_each_list.append(int(each_cat[0]) + 1)
                if each == 'label':
                    aal_ID = Pt_Id + aal_vt_name + each
                    aal_each_list.append(aal_ID)
                    for each_cat in aal_table_dict.items():
                        aal_label = each_cat[1]
                        aal_each_list.append(aal_label)
                if each == 'severity':
                    aal_ID = Pt_Id + aal_vt_name + each
                    aal_each_list.append(aal_ID)
                    for each_cat in aal_table_severity_dict.items():
                        aal_sev = each_cat[1]
                        if aal_sev is not None and each_cat[1] != '':
                            aal_each_list.append(aal_sev)

                if each == 'state':
                    aal_ID = Pt_Id + aal_vt_name + each
                    aal_each_list.append(aal_ID)
                    for each_cat in aal_table_severity_dict.items():
                        aal_sev = each_cat[1]
                        if aal_sev is not None and each_cat[1] != '':
                            if int(aal_sev) != 0:
                                aal_each_list.append('A')
                            else:
                                aal_each_list.append('N')
                class_aal_list.append(aal_each_list)
        # 4-AalValueLimits
        if DC_Data_Type[0:2] == 'AI':
            aal_vl_each_list = []
            min_value = ws_class.cell(row=row_num,
                                              column=class_dict[list_name_dict['lower_limit']] + 1).value
            max_value = ws_class.cell(row=row_num,
                                      column=class_dict[list_name_dict['upper_limit']] + 1).value
            if min_value is None or min_value == '':
                min_value = '-99999996802856925000000000000000000000.000000'
            else:
                min_value = float(min_value) - 0.01
            if max_value is None or max_value == '':
                max_value = '99999996802856925000000000000000000000.000000'
            aal_vl_name = ':aal:V_valueLimits'
            aal_vl_ID = Pt_Id + aal_vl_name
            aal_vl_each_list.append(aal_vl_ID)
            aal_vl_each_list.append(0)
            aal_vl_each_list.append(min_value)
            aal_vl_each_list.append(min_value)
            aal_vl_each_list.append(max_value)
            aal_vl_each_list.append(max_value)
            class_aal_vl_list.append(aal_vl_each_list)

    save_class_excel(class_eq_list, csv_1_eq_path, 'eq')
    xlsx2csv(csv_1_eq_path)
    save_class_excel(class_pt_list, csv_2_pt_path, 'pt')
    xlsx2csv(csv_2_pt_path)
    save_class_excel(class_xac_list, csv_3_xac_path, 'xac')
    xlsx2csv(csv_3_xac_path)
    save_class_excel(class_xal_list, csv_3_xal_path, 'xal')
    xlsx2csv(csv_3_xal_path)
    save_class_excel(class_xfo_list, csv_3_xfo_path, 'xfo')
    xlsx2csv(csv_3_xfo_path)
    save_class_excel(class_dov_list, csv_3_dov_path, 'dov')
    xlsx2csv(csv_3_dov_path)
    save_class_excel(class_dio_list, csv_4_dio_path, 'dioVT')
    xlsx2csv(csv_4_dio_path)
    save_class_excel(class_dal_list, csv_4_dal_path, 'dalVT')
    xlsx2csv(csv_4_dal_path)
    save_class_excel(class_dal10_list, csv_4_dal10_path, 'dalVT10')
    xlsx2csv(csv_4_dal10_path)
    save_class_excel(class_aal_list, csv_4_aal_path, 'aalVT')
    xlsx2csv(csv_4_aal_path)
    save_class_excel(class_aal_vl_list, csv_4_aalvl_path, 'aalVL')
    xlsx2csv(csv_4_aalvl_path)
    save_new_po_csv(new_mapping_dict, alm_save_path)
    save_new_po(alm_content, new_mapping_dict, alm_path.replace('.po', '_new.po'))


def add_list_int_pt():
    #填加点表内部点
    print('请选择需要填加内部点的点表：')
    i = 1
    dict_file = {}
    for each_file in os.listdir(list_path):
        print(str(i) + each_file)
        dict_file[i] = each_file
        i += 1
    select = int(input())
    file_name = list_path + '\\' + dict_file[select]
    file_name1 = file_name + '111'
    wb = openpyxl.load_workbook(file_name, data_only=True)
    ws = wb['IO List']
    max_row_num = ws.max_row
    max_col_num = ws.max_column
    while ws.cell(row=max_row_num, column=9).value is None:
        max_row_num -= 1
    j = 3
    while j <= max_row_num:
        if str(ws['DH' + str(j)].value).upper() == 'TWOSTEPS':
            if str(ws['N' + str(j + 1)].value) == 'DI_INT':
                # 只更新选择内部点地址
                ws['EN' + str(j + 1)].value = ws['EN' + str(j)].value
                ws['EO' + str(j + 1)].value = ws['EO' + str(j)].value
                ws['EP' + str(j + 1)].value = ws['EP' + str(j)].value
                ws['EQ' + str(j + 1)].value = ws['EQ' + str(j)].value
                ws['ES' + str(j + 1)].value = ws['ES' + str(j)].value
                ws['EN' + str(j + 2)].value = ws['EN' + str(j)].value
                ws['EO' + str(j + 2)].value = ws['EO' + str(j)].value
                ws['EP' + str(j + 2)].value = ws['EP' + str(j)].value
                ws['EQ' + str(j + 2)].value = ws['EQ' + str(j)].value
                ws['ES' + str(j + 2)].value = ws['ES' + str(j)].value
                ws['EO' + str(j + 1)].value = str(ws['EO' + str(j)].value).replace('50', '20').replace('51', '21')
                j = j + 1
            else:
                ws.insert_rows(j + 1, 1)
                for each_col in range(1, ws.max_column):
                    ws.cell(row=j + 1, column=each_col).value = ws.cell(row=j, column=each_col).value
                    # ws.cell(row=j + 1, column=each_col).fill = fills.GradientFill(stop=['FF92D050', 'FF92D050'])
                for each_col1 in range(1, ws.max_column):
                    ws.cell(row=j + 1, column=each_col1).fill = PatternFill(fill_type='solid', fgColor="FF92D050")
                    ws.cell(row=j + 1, column=each_col1).font = Font(name='宋体')
                    ws.cell(row=j + 1, column=each_col1).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=j + 1, column=each_col1).border = Border(left=Side(style='thin',color='FF000000'), right=Side(style='thin', color='FF000000'),top=Side(style='thin', color='FF000000'),bottom=Side(style='thin', color='FF000000'))
                ws.insert_rows(j + 2, 1)
                for each_col in range(1, ws.max_column):
                    ws.cell(row=j + 2, column=each_col).value = ws.cell(row=j, column=each_col).value
                for each_col1 in range(1, ws.max_column):
                    ws.cell(row=j + 2, column=each_col1).fill = PatternFill(fill_type='solid', fgColor="FF92D050")
                    ws.cell(row=j + 2, column=each_col1).font = Font(name='宋体')
                    ws.cell(row=j + 2, column=each_col1).alignment = Alignment(horizontal='left', vertical='center')
                    ws.cell(row=j + 2, column=each_col1).border = Border(left=Side(style='thin', color='FF000000'),
                                                                         right=Side(style='thin', color='FF000000'),
                                                                         top=Side(style='thin', color='FF000000'),
                                                                         bottom=Side(style='thin', color='FF000000'))
                ws['DH' + str(j + 1)].value = ''
                ws['DH' + str(j + 2)].value = ''
                ws['N' + str(j + 1)].value = 'DI_INT'
                ws['N' + str(j + 2)].value = 'DO_INT'
                ws['EO' + str(j + 1)].value = str(ws['EO' + str(j)].value).replace('50', '150').replace('51', '151')
                j = j + 3
                max_row_num += 2
        else:
            j = j + 1
    wb.save(file_name)
    wb.close()


def add_rev_reg(wb):
    ws_reg = wb['Revision Register(变更记录)']
    max_row_num = ws_reg.max_row
    while ws_reg.cell(row=max_row_num, column=1).value is None:
        max_row_num -= 1
    tar_row_num = str(max_row_num + 1)
    tar_vis = ''
    for each in str(ws_reg['A' + str(max_row_num)].value).split('.')[0:-1]:
        tar_vis = tar_vis + each + '.'
    tar_num = int(str(ws_reg['A' + str(max_row_num)].value).split('.')[-1])
    tar_num += 1
    tar_vis += str(tar_num)
    cur_date = time.strftime("%Y%m%d", time.localtime())
    ws_reg['A' + tar_row_num].value = tar_vis
    ws_reg['B' + tar_row_num].value = 'Update DB Information(Auto Gen)'
    ws_reg['C' + tar_row_num].value = 'Thales'
    ws_reg['D' + tar_row_num].value = 'Chao'
    ws_reg['E' + tar_row_num].value = time.strftime("%Y/%m/%d", time.localtime())
    return tar_vis, cur_date


def update_dbeq_csv():
    csv_path = cur_path + '\\Generated_csv'
    comp_path = cur_path + '\\Compare_Result\\'
    new_addeqp_file = comp_path + 'new_add_eqp.xlsx'
    up_eqp_file = comp_path + 'up_attr_eqp.xlsx'
    del_eqp_file = comp_path + 'del_eqp.xlsx'
    print('请在此文件列表中进行选择:\n')
    i = 1
    csv_file_dict = {}
    for each_file in os.listdir(csv_path):
        print(str(i) + '. ' + each_file)
        csv_file_dict[i] = each_file
        i += 1
    old_file = csv_path + '\\' + csv_file_dict[int(input('请选择上个版本生成的csv文件:'))]
    cur_file = csv_path + '\\' + csv_file_dict[int(input('请选择本个版本生成的csv文件:'))]
    old_label_dict = {}
    cur_label_dict = {}
    old_eqp_loc_dict = {}
    cur_eqp_loc_dict = {}
    old_ev_list, old_evid_list = read_file_to_list(old_file)
    cur_ev_list, cur_evid_list = read_file_to_list(cur_file)
    old_label_dict, old_eqp_loc_dict = read_eq_attribute_to_list(old_file)
    cur_label_dict, cur_eqp_loc_dict = read_eq_attribute_to_list(cur_file)
    # print(old_evid_list, cur_evid_list)
    # add new equipment
    new_eqp_list = []
    for each in cur_evid_list:
        if each not in old_evid_list:
            for each1 in cur_ev_list:
                if each in each1:
                    add_eqp_list = []
                    for each_ele in each1.split(','):
                        each_ele = each_ele.replace('\n', '')
                        add_eqp_list.append(each_ele)
                    new_eqp_list.append(add_eqp_list)
    save_to_excel(new_eqp_list, new_addeqp_file)
    xlsx2csv(new_addeqp_file)
    # delete equipment
    del_eqp_list = []
    for each in old_evid_list:
        if each not in cur_evid_list:
            del_eqpid_list = []
            del_eqpid_list.append(each)
            del_eqpid_list.append('TRUE')
            del_eqp_list.append(del_eqpid_list)
    save_to_del_excel(del_eqp_list, del_eqp_file)
    xlsx2csv(del_eqp_file)
    # update attribute
    eq_up_list = []
    for each in cur_evid_list:
        if each in old_evid_list:
            up_label_list = []
            old_label = old_label_dict[each]
            cur_label = cur_label_dict[each]
            old_loc = old_eqp_loc_dict[each].replace('\n', '')
            cur_loc = cur_eqp_loc_dict[each].replace('\n', '')
            if old_label != cur_label:
                up_label_list.append(each)
                if cur_label != '':
                    up_label_list.append(cur_label)
                else:
                    up_label_list.append('__ERASE__')
                if old_loc != cur_loc:
                    if cur_loc != '':
                        up_label_list.append(cur_loc)
                    else:
                        up_label_list.append('__ERASE__')
            else:
                if old_loc != cur_loc:
                    up_label_list.append(each)
                    up_label_list.append('')
                    if cur_loc != '':
                        up_label_list.append(cur_loc)
                    else:
                        up_label_list.append('__ERASE__')
            if len(up_label_list):
                eq_up_list.append(up_label_list)
            # print(eq_up_list)
    save_to_update_excel(eq_up_list, up_eqp_file, 'eq')
    xlsx2csv(up_eqp_file)


def update_dbev_csv():
    csv_path = cur_path + '\\Generated_csv'
    comp_path = cur_path + '\\Compare_Result\\'
    new_addeqp_file = comp_path + 'new_add_ev.xlsx'
    del_eqp_file = comp_path + 'del_ev.xlsx'
    up_db_file = comp_path + 'up_deadband_ev.xlsx'
    up_addr_file = comp_path + 'up_address_ev.xlsx'
    print('请在此文件列表中进行选择:\n')
    i = 1
    csv_file_dict = {}
    for each_file in os.listdir(csv_path):
        print(str(i) + '. ' + each_file)
        csv_file_dict[i] = each_file
        i += 1
    old_file = csv_path + '\\' + csv_file_dict[int(input('请选择上个版本生成的csv文件:'))]
    cur_file = csv_path + '\\' + csv_file_dict[int(input('请选择本个版本生成的csv文件:'))]
    old_ev_list, old_evid_list = read_file_to_list(old_file)
    cur_ev_list, cur_evid_list = read_file_to_list(cur_file)

    old_address_dict, old_length_dict, old_deadband_dict, old_scaling_dict, old_varinvalid_dict = read_ev_attribute_to_list(old_file)
    cur_address_dict, cur_length_dict, cur_deadband_dict, cur_scaling_dict, cur_varinvalid_dict = read_ev_attribute_to_list(cur_file)
    new_eqp_list = []
    for each in cur_evid_list:
        if each not in old_evid_list:
            for each1 in cur_ev_list:
                if each in each1:
                    add_eqp_list = []
                    for each_ele in each1.split(','):
                        each_ele = each_ele.replace('\n', '')
                        add_eqp_list.append(each_ele)
                    new_eqp_list.append(add_eqp_list)
    save_ev_to_excel(new_eqp_list, new_addeqp_file)
    xlsx2csv(new_addeqp_file)
    # delete equipment
    del_eqp_list = []
    for each in old_evid_list:
        if each not in cur_evid_list:
            del_eqpid_list = []
            del_eqpid_list.append(each)
            del_eqpid_list.append('TRUE')
            del_eqp_list.append(del_eqpid_list)
    save_to_del_excel(del_eqp_list, del_eqp_file)
    xlsx2csv(del_eqp_file)

    # Update EV Attribute
    ev_up_db_list = []
    ev_up_addr_list = []
    for each in cur_evid_list:
        if each in old_evid_list:
            up_address_list = []
            up_deadband_list = []
            old_address = old_address_dict[each]
            cur_address = cur_address_dict[each]
            old_deadband = old_deadband_dict[each]
            cur_deadband = cur_deadband_dict[each]
            old_scaling = old_scaling_dict[each]
            cur_scaling = cur_scaling_dict[each]
            old_varinvalid = old_varinvalid_dict[each]
            cur_varinvalid = cur_varinvalid_dict[each]
            if old_scaling != cur_scaling:
                up_deadband_list.append(each)
                if cur_scaling != '':
                    up_deadband_list.append(cur_scaling)
                else:
                    up_deadband_list.append('__ERASE__')
                if old_deadband != cur_deadband:
                    if cur_deadband != '':
                        up_deadband_list.append(cur_deadband)
                    else:
                        up_deadband_list.append('__ERASE__')
            else:
                if old_deadband != cur_deadband:
                    up_deadband_list.append(each)
                    up_deadband_list.append('')
                    if cur_deadband != '':
                        up_deadband_list.append(cur_deadband)
                    else:
                        up_deadband_list.append('__ERASE__')
            if len(up_deadband_list):
                ev_up_db_list.append(up_deadband_list)
            if old_address != cur_address:
                up_address_list.append(each)
                if cur_address != '':
                    up_address_list.append(cur_address)
                else:
                    up_address_list.append('__ERASE__')
            if old_varinvalid != cur_varinvalid:
                up_address_list.append(each)
                up_address_list.append('')
                if cur_varinvalid != '':
                    up_address_list.append(cur_varinvalid)
                else:
                    up_address_list.append('__ERASE__')
            if len(up_address_list):
                ev_up_addr_list.append(up_address_list)
    save_to_update_excel(ev_up_db_list, up_db_file, 'db')
    xlsx2csv(up_db_file)
    save_to_update_excel(ev_up_addr_list, up_addr_file, 'addr')
    xlsx2csv(up_addr_file)


def read_file_to_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as f1:
        list1 = f1.readlines()
        list_id = []
        for i in range(1, len(list1)):
            list_id.append(list1[i].split(',')[0])
        list1 = list1[1:]
    return list1, list_id


def read_eq_attribute_to_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as f1:
        list1 = f1.readlines()
        list_id = []
        label_dict = {}
        eqp_loc_dict = {}
        for i in range(1, len(list1)):
            list_id.append(list1[i].split(',')[0])
            label_dict[list1[i].split(',')[0]] = list1[i].split(',')[2]
            eqp_loc_dict[list1[i].split(',')[0]] = list1[i].split(',')[8]
        list1 = list1[1:]
        # print(label_dict, eqp_loc_dict)
    return label_dict, eqp_loc_dict


def read_ev_attribute_to_list(file_path):
    with open(file_path, 'r', encoding='utf-8') as f1:
        list1 = f1.readlines()
        list_id = []
        address_dict = {}
        length_dict = {}
        deadband_dict = {}
        scaling_dict = {}
        varinvalid_dict = {}
        for i in range(1, len(list1)):
            list_id.append(list1[i].split(',')[0])
            address_dict[list1[i].split(',')[0]] = list1[i].split(',')[2]
            length_dict[list1[i].split(',')[0]] = list1[i].split(',')[4]
            deadband_dict[list1[i].split(',')[0]] = list1[i].split(',')[5]
            scaling_dict[list1[i].split(',')[0]] = list1[i].split(',')[9]
            varinvalid_dict[list1[i].split(',')[0]] = list1[i].split(',')[8]
        list1 = list1[1:]
    return address_dict, length_dict, deadband_dict, scaling_dict, varinvalid_dict


def xlsx2csv(filename):
    try:
        xlsx_file_reader = load_workbook(filename=filename)
        for sheet in xlsx_file_reader.get_sheet_names():
            # 每个sheet输出到一个csv文件中，文件名用xlsx文件名和sheet名用'_'连接
            csv_filename = '{xlsx}.csv'.format(
                xlsx=os.path.splitext(filename.replace(' ', '_'))[0],
                sheet=sheet.replace(' ', '_'))

            with open(csv_filename, 'w', newline='') as f_out:
                csv_file_writer = csv.writer(f_out)

                sheet_ranges = xlsx_file_reader[sheet]
                for row in sheet_ranges.rows:
                    row_container = []
                    for cell in row:
                        if cell.value is not None:
                            if type(cell.value) == str:
                                row_container.append(str(cell.value))
                            else:
                                row_container.append(str(cell.value))
                        else:
                            cell.value = ''
                            row_container.append(str(cell.value))
                    # print(row_container)
                    csv_file_writer.writerow(row_container)
        os.remove(filename)
    except Exception as e:
        print(e)


def replace_alm_po(old_alm_path, new_alm_path):
    cur_dat = time.strftime("%Y%m%d%H%M%S", time.localtime())
    if os.path.exists(new_alm_path):
        os.rename(old_alm_path, old_alm_path.replace('.po', '_' + str(cur_dat) + '.po'))
        shutil.copyfile(new_alm_path, old_alm_path)


def file_check_diff(old_file, new_file):
    diff = 1
    if os.path.exists(new_file):
        if filecmp.cmp(old_file, new_file):
            print("alm_cn.po文件无更新")
        else:
            print("alm_cn.po文件有更新，请确认是否需要替换为新的文件!")
            diff = 0
    else:
        print("No alm_cn_new.po file in the folder")
        diff = 0
    return diff


def insert_dup_value(new_mapping_dict,alm_dict, attr, tmp_res):
    i = 0
    num = 1
    org_tem_res = tmp_res
    while num == 1:
        try:
            alm_dict[attr] = tmp_res
            new_mapping_dict[attr] = tmp_res
            num = 0
        except ValueDuplicationError as e:
            tmp_res = org_tem_res
            i += 1
            tmp_res = tmp_res + str(i)
            num = 1
    return tmp_res, new_mapping_dict, alm_dict


def fep_com_state(swc_id):
    for item, value in fep_com_dict.items():
        if str(value[0]) == str(swc_id):
            com_state = item
    return com_state


def gen_com_ev(sta_name):
    fep_ve_list = []
    target_dict = {}
    if sta_name != 'OCC':
        target_dict = fep_com_sta_dict
    else:
        target_dict = fep_com_dict
    for item, value in target_dict.items():
        fep_grp_list = []
        fep_grp_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item))
        fep_grp_list.append(str(item))
        fep_grp_list.append('')
        fep_grp_list.append('')
        fep_grp_list.append('')
        fep_grp_list.append('')
        fep_grp_list.append(str(item))
        fep_grp_list.append('GRP_VE')
        fep_ve_list.append(fep_grp_list)
        fep_comstatus_list = []
        fep_comstatus_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0001')
        fep_comstatus_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-COM')
        fep_comstatus_list.append('$ComStatus-00' + "%(a)03d" % {'a': value[1]} + '/4+5')
        fep_comstatus_list.append('DEIV')
        fep_comstatus_list.append('4')
        fep_comstatus_list.append('0')
        fep_comstatus_list.append(str(item) + ' - FEP COMM STATUS')
        fep_comstatus_list.append('VE')
        fep_ve_list.append(fep_comstatus_list)
        fep_infstatus_list = []
        fep_infstatus_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0002')
        fep_infstatus_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-INF')
        fep_infstatus_list.append('$InfStatus-00' + "%(a)03d" % {'a': value[0]} + '/4+5')
        fep_infstatus_list.append('DEIV')
        fep_infstatus_list.append('4')
        fep_infstatus_list.append('0')
        fep_infstatus_list.append(str(item) + ' - FEP INF STATUS')
        fep_infstatus_list.append('VE')
        fep_ve_list.append(fep_infstatus_list)
        fep_setcom_list = []
        fep_setcom_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0003')
        fep_setcom_list.append('deo' + sta_name + 'SYSFEP' + str(item) + '-SET')
        fep_setcom_list.append('$SetComStatus-00' + "%(a)03d" % {'a': value[0]} + '-1')
        fep_setcom_list.append('DEOV')
        fep_setcom_list.append('4')
        fep_setcom_list.append('0')
        fep_setcom_list.append(str(item) + ' - FEP SET COM STATUS')
        fep_setcom_list.append('VE')
        fep_ve_list.append(fep_setcom_list)
        fep_link1_list = []
        fep_link1_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0004')
        fep_link1_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-LINK1')
        fep_link1_list.append('000/001/00' + "%(a)03d" % {'a': value[1]} + '/4/000001')
        fep_link1_list.append('DEIV')
        fep_link1_list.append('4')
        fep_link1_list.append('0')
        fep_link1_list.append(str(item) + ' - FEP LINK1 STATUS')
        fep_link1_list.append('VE')
        fep_ve_list.append(fep_link1_list)
        fep_link2_list = []
        fep_link2_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0005')
        fep_link2_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-LINK2')
        fep_link2_list.append('000/001/00' + "%(a)03d" % {'a': value[1]} + '/5/000001')
        fep_link2_list.append('DEIV')
        fep_link2_list.append('4')
        fep_link2_list.append('0')
        fep_link2_list.append(str(item) + ' - FEP LINK2 STATUS')
        fep_link2_list.append('VE')
        fep_ve_list.append(fep_link2_list)
        fep_link3_list = []
        fep_link3_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0006')
        fep_link3_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-LINK3')
        fep_link3_list.append('000/002/00' + "%(a)03d" % {'a': value[1]} + '/4/000001')
        fep_link3_list.append('DEIV')
        fep_link3_list.append('4')
        fep_link3_list.append('0')
        fep_link3_list.append(str(item) + ' - FEP LINK3 STATUS')
        fep_link3_list.append('VE')
        fep_ve_list.append(fep_link3_list)
        fep_link4_list = []
        fep_link4_list.append(':R:A:POLE_' + sta_name + ':FEP:SYS:' + str(item) + ':VE0007')
        fep_link4_list.append('dei' + sta_name + 'SYSFEP' + str(item) + '-LINK4')
        fep_link4_list.append('000/002/00' + "%(a)03d" % {'a': value[1]} + '/5/000001')
        fep_link4_list.append('DEIV')
        fep_link4_list.append('4')
        fep_link4_list.append('0')
        fep_link4_list.append(str(item) + ' - FEP LINK4 STATUS')
        fep_link4_list.append('VE')
        fep_ve_list.append(fep_link4_list)
    return fep_ve_list


def update_po_to_db_machine():
    server_alm_po = '\\\\192.168.1.200\\d\\NBL3_ALMPO\\alm_cn.po'
    old_alm_path = alm_folder + 'alm_cn.po'
    shutil.copyfile(old_alm_path, server_alm_po)


def file_check_exist():
    class_file_list = os.listdir(class_path)
    error_code = 0
    list_file_list = os.listdir(list_path)
    print("请确认以下文件:")
    if len(class_file_list) > 1:
        print("请确认在class文件夹中只能有一个class文件!")
        error_code = 1
    elif len(class_file_list) < 1:
        print("请在Class 文件夹中添加IO Class!")
        error_code = 1
    else:
        print("IO Class : " + class_file_list[0])
    if len(list_file_list) == 0:
        print("请在List 文件夹中添加对应的IO List!")
        error_code = 1
    else:
        print("IO List:")
        for each_file in list_file_list:
            print(each_file + '\n')
    return error_code


if __name__ == "__main__":
    old_alm_path = alm_folder + 'alm_cn.po'
    new_alm_path = alm_folder + 'alm_cn_new.po'
    server_alm_po = '\\\\192.168.1.200\\d\\NBL3_ALMPO\\alm_cn.po'
    error_code = file_check_exist()
    if error_code == 1:
        time.sleep(10)
        exit("请检查文件!")
    alm_code = file_check_diff(old_alm_path, server_alm_po)
    if alm_code == 0:
        res = input('是否需要从DB服务器上获取并更新alm_cn.po文件: y/n(直接回车默认为y)\n')
        if res == 'y' or res == '':
            replace_alm_po(old_alm_path, server_alm_po)
        cmd = str(input("请对比DB服务器上，确认alm_cn.po文件为最新，按 y (或直接回车)继续:"))
        if cmd == '' or cmd.upper() == 'Y':
            pass
        else:
            print("准备退出程序，请检查所需文件................")
            time.sleep(5)
            exit()
    file_check_diff(old_alm_path, server_alm_po)
    print('请选择需要的操作,输入Q退出：')
    print('1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
    choice = input()
    while choice != 'Q':
        if choice == '1':
            fill_list()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '2':
            file_check_diff(old_alm_path, new_alm_path)
            res = input('是否需要更新alm_cn.po文件: y/n(直接回车默认为y)\n')
            if res == 'y' or res == '':
                replace_alm_po(old_alm_path, new_alm_path)
            gen_csv()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '3':
            file_check_diff(old_alm_path, new_alm_path)
            res = input('是否需要更新alm_cn.po文件: y/n(直接回车默认为y)\n')
            if res == 'y' or res == '':
                replace_alm_po(old_alm_path, new_alm_path)
            gen_class_csv()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '4':
            add_list_int_pt()
            print('请等待操作完成................')
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '5':
            update_dbeq_csv()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '6':
            update_dbev_csv()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '7':
            sta_name = input('请输入需要生成com ev的车站名:\n')
            cur_path = os.getcwd()
            out_path = cur_path + '\\Fep_Com'
            ev_save_path = out_path + '\\FEP_COM_EV_' + str(sta_name) + '.xlsx'
            save_ev_to_excel(gen_com_ev(sta_name), ev_save_path)
            xlsx2csv(ev_save_path)
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print('1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        elif choice == '8':
            file_check_diff(old_alm_path, new_alm_path)
            res = input('是否需要更新alm_cn.po文件: y/n(直接回车默认为y)\n')
            if res == 'y' or res == '':
                replace_alm_po(old_alm_path, new_alm_path)
            update_po_to_db_machine()
            print('生成完成！')
            print('请选择需要的操作,输入Q退出：')
            print(
                '1. 自动在点表中添加信息.\n2. 根据点表生成csv文件.\n3. 根据类表生成csv文件\n4. 点表中添加内部点\n5. 生成前后版本更新equipment.csv\n6. 生成前后版本更新ev.csv\n7. 生成车站fep com EV文件\n8. 上传alm_cn.po至DB Server')
            choice = input()
        else:
            print('输入有误，请重新输入')
            choice = input('请选择需要的操作,输入Q退出：')
