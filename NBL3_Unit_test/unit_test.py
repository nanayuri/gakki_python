import difflib
import paramiko, datetime, os, time
import openpyxl
from openpyxl import load_workbook

hostname = '192.168.1.41'
username = 'nbl2'
password = 'nbl2iscs'
port = 22
local_dir = os.getcwd()
package_version = 'OCC_A1TEMP2_181026'
remote_dir = '/export/home/nbl2/' + package_version + '/log/'
ut_dir = remote_dir + 'unit_test/'
script_dir = local_dir + '/script'
io_dir = local_dir + '/io_list'

station_dict = {
    '高塘桥站': 'GTQ',
    '句章路站': 'JZL',
    '鄞州客运总站 ': 'YZC',
    '南部商务区站': 'SBD',
    '鄞州区政府站': 'YZG',
    '四明中路站': 'SMZ',
    '锦寓路站': 'JYL',
    '钱湖北路站': 'QHB',
    '仇毕站': 'QBZ',
    '儿童公园站': 'ETZ',
    '樱花公园站': 'YHZ',
    '体育馆站': 'TYG',
    '明楼站': 'MLZ',
    '中兴大桥南站': 'ZXN',
    '大通桥站': 'DTZ',
    '姜山站': 'JSZ',
    '狮山站': 'SSZ',
    '明辉路站': 'MHL',
    '朝阳站': 'CYZ',
    '方桥站': 'FQZ',
    '琎琳站': 'JLZ',
    '南渡站': 'NDZ',
    '大成东路站': 'DCZ',
    '金海路站': 'JHL',
    '首南车辆段': 'DEP',
    '奉化停车场': 'FHT',
    '控制中心': 'OCC',
    '樱花主变': 'YHB',
    '双桥主变': 'SQB',
    '蒋家主变电站': 'JJB'
}


class ExcValue:
    def __init__(self, file, io_type):
        self.wb = openpyxl.load_workbook(file, data_only=True)
        if io_type == 'class':
            self.sheet = self.wb['CLASS']
        elif io_type == 'list':
            self.sheet = self.wb['IO List']

    def value(self, x, y):
        return self.sheet.cell(row=x, column=y).value

    def getvalue(self, x, y):
        return self.sheet[str(x) + str(y)].value

    def cell(self, x, y):
        return self.sheet.cell(row=x, column=y)

    def write_value(self, x, y, val):
        self.sheet.cell(row=x, column=y).value = val

    def get_col_num(self):
        return self.sheet.columns

    def max_row(self):
        return self.sheet.max_row

    def save(self, files_name):
        self.wb.save(files_name)


def compare_file(file1, file2):
    with open(file1, 'r', encoding='utf-8') as f1:
        text1_lines = f1.readlines()
    with open(file2, 'r', encoding='utf-8') as f2:
        text2_lines = f2.readlines()
    d = difflib.Differ()
    diff = d.compare(text1_lines, text2_lines)
    out_list = []
    for each_line in list(diff):
        if each_line[0] == '+' and each_line[2] == '-':
            print(each_line)
            out_list.append(each_line)
    return out_list


def upload_file(files, dir):
    try:
        t = paramiko.Transport((hostname, port))
        t.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(t)
        content = 'cd  ' + remote_dir + ';mkdir unit_test'
        add_right = 'cd ' + ut_dir + '; chmod +x *.*'
        run_script(content)
        for f in files:
            sftp.put(os.path.join(dir, f), os.path.join(ut_dir, f))
        run_script(add_right)
        t.close()
    except Exception:
        print('connect error!')


def download_file(dir):
    try:
        t = paramiko.Transport((hostname, port))
        t.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(t)
        content = 'cd ' + ut_dir
        run_script(content)
        sftp.get(ut_dir + 'old_event.txt', dir + '\\old_event.txt')
        sftp.get(ut_dir + 'new_event.txt', dir + '\\new_event.txt')
        t.close()
    except Exception:
        print('connect error!')


def run_script(content):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=hostname, port=port, username=username, password=password)
    stdin, stdout, stderr = ssh.exec_command(content)
    print(stdout.readlines())
    ssh.close()


def get_list_value():
    pass


def gen_txt(file_name):
    total_list = []
    io_list = ExcValue(file_name, 'list')
    max_row_num = io_list.max_row()
    while io_list.value(max_row_num, 9) is None:
        max_row_num -= 1
    for row_num in range(3, max_row_num + 1):
        station_code = io_list.value(row_num, 5)
        system = io_list.value(row_num, 7)
        eqpt_loc = io_list.value(row_num, 6)
        eqpt_code = io_list.value(row_num, 9)
        eqpt_desc = io_list.value(row_num, 10)
        eqpt_ident = io_list.value(row_num, 11)
        eqpt_attr = io_list.value(row_num, 12)
        pt_type = io_list.value(row_num, 14)
        v0_label = io_list.value(row_num, 16)
        v0_serv = io_list.value(row_num, 17)
        v1_label = io_list.value(row_num, 20)
        v1_serv = io_list.value(row_num, 21)
        v2_label = io_list.value(row_num, 24)
        v2_serv = io_list.value(row_num, 25)
        v3_label = io_list.value(row_num, 28)
        v3_serv = io_list.value(row_num, 29)
        v4_label = io_list.value(row_num, 32)
        v4_serv = io_list.value(row_num, 33)
        v5_label = io_list.value(row_num, 36)
        v5_serv = io_list.value(row_num, 37)
        v6_label = io_list.value(row_num, 40)
        v6_serv = io_list.value(row_num, 41)
        v7_label = io_list.value(row_num, 44)
        v7_serv = io_list.value(row_num, 45)
        pt_name = io_list.value(row_num, 123)
        pt_alias = io_list.value(row_num, 125)
        label_list = [v0_label, v1_label, v2_label, v3_label, v4_label, v5_label, v6_label, v7_label]
        bit_dict = {v0_label: v0_serv,
                    v1_label: v1_serv,
                    v2_label: v2_serv,
                    v3_label: v3_serv,
                    v4_label: v4_serv,
                    v5_label: v5_serv,
                    v6_label: v6_serv,
                    v7_label: v7_serv
                    }
        bit_dict1 = {v0_label: 0,
                    v1_label: 1,
                    v2_label: 2,
                    v3_label: 3,
                    v4_label: 4,
                    v5_label: 5,
                    v6_label: 6,
                    v7_label: 7
                    }
        for each_label in label_list:
            if each_label is not None:
                target_list = []
                for each in [pt_alias, pt_name, pt_type, station_dict[station_code], system, eqpt_loc, eqpt_code, eqpt_desc,
                                   eqpt_ident, eqpt_attr, each_label, str(bit_dict1[each_label]), str(bit_dict[each_label])]:
                    target_list.append(str(each) + ',')
                target_list.append('\n')

                total_list.append(target_list)

        save_to_txt(file_name, total_list)


def save_to_txt(file_name, content):
    with open(str(file_name).replace('.xlsx', '.txt'), 'w', encoding='utf-8') as f1:
        for each_line in content:
            f1.writelines(each_line)


if __name__ == "__main__":
    final_list = []
    # Step 1. IO list 转换为 txt 文档， 需要的属性有：车站，设备描述，设备标签，属性描述，报警事件点标签，报警严重度
    file_temp = io_dir + '\\' + 'NBL3 NF IO List AFC YHZ(樱花公园站) V1.2 (20181011).xlsx'
    #gen_txt(file_temp)
    print('-----IO list.txt for ' + file_temp + ' has been created Successfully! -----')

    # Step 2. 上传文件到Server中
    upload_file(['save_new.sh', 'save_old.sh'])
    print('-----.sh script files are uploaded to server successfully! -----')
    # Step 3.在Server中，执行.py文件去选取每一个点
    # command = 'cd /export/home/nbl2/OCC_A1TEMP2_181023/log/unit_test; ./abc.py'
    # run_script(command)
    # 首先set 通讯状态点
    file_gen = str(file_temp).replace('.xlsx', '.txt')
    with open(file_gen, 'r', encoding='utf-8') as f1:
        io_cont = f1.readlines()
    station_name = io_cont[0].split(',')[3]
    sys_name = io_cont[0].split(',')[4]
    com_file_template = 'set_com.sh_template'
    com_file = 'set_com.sh'
    with open(com_file_template, 'r') as f_com:
        com_cont = f_com.read()
    with open(com_file, 'w', newline='\n') as f_com_new:
        f_com_new.write(str(com_cont).replace('station_name', station_name).replace('sys_name', sys_name))
    upload_file(['set_com.sh', 'comstate.tcl'])
    command = 'cd ' + ut_dir + ';./set_com.sh'
    run_script(command)
    # print('-----Com State point setting operation' + str(com_cont).replace('station_name', station_name).replace('sys_name', sys_name) + ' has been created Successfully! -----')

    for each_line in io_cont:
        print(each_line)
        # Step 4. 保存eventlist到old_event.txt
        # command1 = 'cd ' + ut_dir + ';rm old_event.txt'
        # run_script(command1)
        command = 'cd ' + ut_dir + ';./save_old.sh'
        print(command)
        run_script(command)
        print('-----Original status for eventlist has been saved successfully! -----')

        # Step 5. 执行tcl去模拟值的变化
        each_list = each_line.split(',')
        target_pt = each_list[0].replace('<alias>', '') + ':' + each_list[1]
        target_value = each_list[-3]
        # 生成对应.sh文件
        sh_file = local_dir + '\\' + 'set_pt.sh_template'
        tar_sh_file = local_dir + '\\' + 'set_pt.sh'
        with open(sh_file, 'r') as f2:
            sh_cont = f2.read()
        with open(tar_sh_file, 'w', newline='\n') as f3:
            f3.write(str(sh_cont).replace('replace_pt_alias', target_pt).replace('replace_value', target_value))
        # Step 6. 模拟点的值
        upload_file(['pt_sim.tcl', 'set_pt.sh'])
        command = 'cd ' + ut_dir + ';./set_pt.sh'
        run_script(command)
        # print('-----The point' + str(sh_cont).replace('replace_pt_alias', target_pt) + ' successfully! -----')
        # Step 7. 保存eventlist到new_event.txt
        command = 'cd ' + ut_dir + ';./save_new.sh'
        run_script(command)
        print('-----New status for eventlist has been saved successfully! -----')
        # Step 8. 上传old_event.txt和new_event.txt
        download_file()
        print('-----Old and new event list are downloaded from server successfully! -----')
        # Step 9. 比较前后两份eventlist
        compare_list = compare_file('old_event.txt', 'new_event.txt')
        # Step 10. 比较模拟的结果和io中的信息
        # 首先处理compare_list
        final_list.append(compare_list)

        if len(compare_list) != 0:
            main_list = compare_list[0].split('"')[5]
            alias_new = '<alias>' + str(main_list[13]).replace(':', '')
            alias_old = str(each_line[0]) + str(each_line[1])

    with open('out.txt', 'w') as f_out:
        for each_line in final_list:
            f_out.writelines(each_line)
    '''
    
    #Step 5. 执行tcl去模拟值的变化
    command = 'cd /export/home/nbl2/OCC_A1TEMP2_181023/log/unit_test; ./test.sh ' + station + ' ' + system
    run_script(command)
    #Step 6. 保存eventlist到new_event.txt
    command = 'cd /export/home/nbl2/OCC_A1TEMP2_181023/log/unit_test; scsolsshow -NAlmServer -lEventList -r >  new_event.txt'
    run_script(command)
    #Step 6.1 取回old_event.txt和new_event.txt
    download_file()
    #Step 7. 比较old_event.txt和new_event.txt
    compare_file('old_event.txt', 'new_event.txt')
    #Step 8. 将比较出的文件与Step 3中的每一个点的信息做比较
    diff_check()
    '''