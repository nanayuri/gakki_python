import difflib
import paramiko, datetime, os, time, sys
import openpyxl
import Para as p

with open('Para.py', 'r') as f:
    para_list = f.readlines()
    print(para_list)
hostname = para_list[0].split("'")[1]
username = para_list[1].split("'")[1]
password = para_list[2].split("'")[1]
port = int(para_list[3].split("'")[1])
local_dir = os.getcwd()
package_version = para_list[4].split("'")[1]
print(hostname, username, password, port, package_version)
remote_dir = '/export/home/' + username + '/' + package_version + '/log/'
ut_dir = remote_dir + 'unit_test/'
script_dir = local_dir + '\script'
io_dir = local_dir + '\io_list'

station_dict = {
    '高塘桥站': 'GTQ',
    '句章路站': 'JZL',
    '鄞州客运总站': 'YZC',
    '南部商务区站': 'SBD',
    '鄞州区政府站': 'YZG',
    '四明中路站': 'SMZ',
    '锦寓路站': 'JYL',
    '钱湖北路站': 'QHB',
    '仇毕站': 'QBZ',
    '儿童公园站': 'ETZ',
    '樱花公园站': 'YHZ',
    '体育馆站': 'TYG',
    '明楼站': 'MLZ',
    '中兴大桥南站': 'ZXN',
    '大通桥站': 'DTZ',
    '姜山站': 'JSZ',
    '狮山站': 'SSZ',
    '明辉路站': 'MHL',
    '朝阳站': 'CYZ',
    '方桥站': 'FQZ',
    '琎琳站': 'JLZ',
    '南渡站': 'NDZ',
    '大成东路站': 'DCZ',
    '金海路站': 'JHL',
    '首南车辆段': 'DEP',
    '奉化停车场': 'FHT',
    '控制中心': 'OCC',
    '樱花主变': 'YHB',
    '双桥主变': 'SQB',
    '蒋家主变电站': 'JJB'
}
occ_dict = {
    'PSCADA': 'OCC',
    'AFC': 'OCC',
    'MPSCADA': 'OCC',
    'SIG': 'OCC',
    'PSD': 'OCC',
    'AFC': 'OCC',
    'BAS': 'OCCB',
    'UPS': 'OCCB',
    'FG': 'OCCB',
    'DTS': 'OCCB',
    'ASD': 'OCCB',
    'FAS': 'OCCC',
    'ALM': 'OCCC',
    'ACS': 'OCCC',
    'PA': 'OCCC',
    'PIS': 'OCCC',
    'CCTV': 'OCCC'
}

class ExcValue:
    def __init__(self, file, io_type):
        self.wb = openpyxl.load_workbook(file, data_only=True)
        if io_type == 'class':
            self.sheet = self.wb['CLASS']
        elif io_type == 'list':
            self.sheet = self.wb['IO List']

    def value(self, x, y):
        return self.sheet.cell(row=x, column=y).value

    def getvalue(self, x, y):
        return self.sheet[str(x) + str(y)].value

    def cell(self, x, y):
        return self.sheet.cell(row=x, column=y)

    def write_value(self, x, y, val):
        self.sheet.cell(row=x, column=y).value = val

    def get_col_num(self):
        return self.sheet.columns

    def max_row(self):
        return self.sheet.max_row

    def save(self, files_name):
        self.wb.save(files_name)


def compare_file(file1, file2):
    with open(file1, 'r', encoding='utf-8') as f1:
        text1_lines = f1.readlines()
    with open(file2, 'r', encoding='utf-8') as f2:
        text2_lines = f2.readlines()
    d = difflib.Differ()
    diff = d.compare(text1_lines, text2_lines)
    out_list = []
    for each_line in list(diff):
        if each_line[0] == '+' and each_line[2] == '-':
            print(each_line)
            out_list.append(each_line)
    return out_list


def upload_file(files, dir):
    try:
        t = paramiko.Transport((hostname, port))
        t.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(t)
        content = 'cd  ' + remote_dir + ';mkdir unit_test'
        add_right = 'cd ' + ut_dir + '; chmod +x *.*'
        run_script(content)
        for f in files:
            print(os.path.join(dir, f))
            sftp.put(os.path.join(dir, f), os.path.join(ut_dir, f))
        run_script(add_right)
        t.close()
    except Exception:
        print('connect error!')


def download_file(file_name, dir):
    try:
        t = paramiko.Transport((hostname, port))
        t.connect(username=username, password=password)
        sftp = paramiko.SFTPClient.from_transport(t)
        content = 'cd ' + ut_dir
        run_script(content)
        print(dir + '\\' + file_name)
        sftp.get(ut_dir + file_name, dir + '\\' + file_name)
        # sftp.get(ut_dir + 'new_event.txt', local_dir + '\\new_event.txt')
        t.close()
    except Exception:
        print('connect error!')


def run_script(content):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(hostname=hostname, port=port, username=username, password=password)
    stdin, stdout, stderr = ssh.exec_command(content)
    for each in stdout.readlines():
        print(each)
    ssh.close()


def get_list_value():
    pass


def gen_txt(file_name):
    total_list = []
    io_list = ExcValue(file_name, 'list')
    max_row_num = io_list.max_row()
    while io_list.value(max_row_num, 9) is None:
        max_row_num -= 1
    check_num_dict = {}
    for row_num in range(3, max_row_num + 1):
        station_code = io_list.value(row_num, 5)
        system = io_list.value(row_num, 7)
        eqpt_loc = io_list.value(row_num, 6)
        eqpt_code = io_list.value(row_num, 9)
        eqpt_desc = io_list.value(row_num, 10)
        eqpt_ident = io_list.value(row_num, 11)
        eqpt_attr = io_list.value(row_num, 12)
        pt_type = io_list.value(row_num, 14)
        v0_label = io_list.value(row_num, 16)
        v0_serv = io_list.value(row_num, 17)
        v1_label = io_list.value(row_num, 20)
        v1_serv = io_list.value(row_num, 21)
        v2_label = io_list.value(row_num, 24)
        v2_serv = io_list.value(row_num, 25)
        v3_label = io_list.value(row_num, 28)
        v3_serv = io_list.value(row_num, 29)
        v4_label = io_list.value(row_num, 32)
        v4_serv = io_list.value(row_num, 33)
        v5_label = io_list.value(row_num, 36)
        v5_serv = io_list.value(row_num, 37)
        v6_label = io_list.value(row_num, 40)
        v6_serv = io_list.value(row_num, 41)
        v7_label = io_list.value(row_num, 44)
        v7_serv = io_list.value(row_num, 45)
        pt_name = io_list.value(row_num, 123)
        pt_alias = io_list.value(row_num, 125)
        if v0_label == v1_label:
            v1_label = str(v1_label) + '_dup'
        if v0_label == v3_label:
            v3_label = str(v3_label) + '_dup3'
        if v0_label == v2_label:
            v2_label = str(v2_label) + '_dup2'
        if v1_label == v2_label:
            v2_label = str(v2_label) + '_dup12'
        if v1_label == v3_label:
            v3_label = str(v3_label) + '_dup13'
        if v2_label == v3_label:
            v3_label = str(v3_label) + '_dup23'
        label_list = [v0_label, v1_label, v2_label, v3_label, v4_label, v5_label, v6_label, v7_label]
        bit_dict = {v0_label: v0_serv,
                    v1_label: v1_serv,
                    v2_label: v2_serv,
                    v3_label: v3_serv,
                    v4_label: v4_serv,
                    v5_label: v5_serv,
                    v6_label: v6_serv,
                    v7_label: v7_serv
                    }
        bit_dict1 = {v0_label: 0,
                     v1_label: 1,
                     v2_label: 2,
                     v3_label: 3,
                     v4_label: 4,
                     v5_label: 5,
                     v6_label: 6,
                     v7_label: 7
                     }
        if io_list.cell(row_num, 5).font.strike is True:
            continue
        if eqpt_code != 'IBST':
            if str(pt_type)[0:2] == 'DI' and str(pt_type) != 'DI_RESP' and eqpt_code[0:3] != 'STA':
                for each_label in label_list:
                    if each_label is not None and str(each_label) != ' ':
                        target_list = []
                        for each in [pt_alias, pt_name, pt_type, station_dict[station_code], system, eqpt_loc, eqpt_code,
                                     eqpt_desc,
                                     eqpt_ident, eqpt_attr, str(each_label).replace('_dup', '').replace('_dup3', '').replace('_dup2', '').replace('_dup13', '').replace('_dup23', '').replace('_dup12', ''), str(bit_dict1[each_label]), str(bit_dict[each_label])]:
                            target_list.append(str(each) + ',')
                        target_list.append('\n')

                        total_list.append(target_list)
                if pt_alias not in check_num_dict:
                    check_num_dict[pt_alias] = set([pt_name])
                else:
                    check_num_dict[pt_alias].add(pt_name)
            else:
                if pt_alias not in check_num_dict:
                    check_num_dict[pt_alias] = set([pt_name])
                else:
                    check_num_dict[pt_alias].add(pt_name)

        txt_name = save_to_txt(file_name, total_list)
    # print(check_num_dict)
    return txt_name, check_num_dict


def save_to_txt(file_name, content):
    with open(str(file_name).replace('.xlsx', '.txt'), 'w', encoding='utf-8') as f1:
        for each_line in content:
            f1.writelines(each_line)
    return str(file_name).replace('.xlsx', '.txt')


def window_to_unix():
    with open('run_test.py', 'r', encoding='utf-8') as f_run:
        content = f_run.read()
    with open(script_dir + '\\run_test_new.py', 'w', newline='\n', encoding='utf-8') as f_run_new:
        f_run_new.write(content)


if __name__ == "__main__":

    error_code = 0
    # Step 1. IO list 转换为 txt 文档， 需要的属性有：车站，设备描述，设备标签，属性描述，报警事件点标签，报警严重度
    error_list = []
    run_script('cd ' + ut_dir + ';rm -rf *.txt;rm -rf *.tcl;rm -rf *.sh*')
    for i in os.listdir("."):
        if os.path.splitext(i)[-1] == ".txt" or os.path.splitext(i)[-1] == ".tcl":
            os.remove(i)
    for i in os.listdir(io_dir + '\\' + "."):
        #if os.path.splitext(i)[-1] == ".txt" or os.path.splitext(i)[-1] == ".tcl":
            #os.remove(i)
        if os.path.splitext(i)[-1] == ".xlsx" and os.path.splitext(i)[0][0] != "~":
            file_temp = io_dir + '\\' + str(i)
            print(file_temp)
            txt_name = gen_txt(file_temp)[0].split('\\')[-1]
            check_dict = gen_txt(file_temp)[1]
            if txt_name.split(' ')[1] != 'NF':
                sys_name = txt_name.split(' ')[3]
            else:
                sys_name = txt_name.split(' ')[4]

            occ_name = occ_dict[sys_name]
            get_alias_list = []
            get_alias_list.append('#!/bin/tclsh\n')
            get_alias_list.append('ScsDbm::init\n')
            get_alias_list.append('set fp [open "get_alias.txt" w]\n')
            for each in check_dict.keys():
                tcl_sent = 'puts $fp [::ScsDbm::getChildrenAliases ' + occ_name + ' ' + str(each) + ']\n'
                get_alias_list.append(tcl_sent)
            get_alias_list.append('close $fp\n')
            get_alias_name = txt_name.replace('.txt', '_check_pt.tcl')
            with open(get_alias_name, 'w', newline='\n') as f_ckpt:
                f_ckpt.writelines(get_alias_list)
            with open(script_dir + '/get_child.sh_template', 'r') as f_child:
                child_cont = f_child.read().replace('file_alias', txt_name.replace('.txt', '_check_pt.tcl'))
            with open(script_dir + '/get_child.sh', 'w', encoding='utf-8', newline='\n') as f_ch:
                f_ch.write(child_cont)

            upload_file([get_alias_name], local_dir)
            upload_file(['get_child.sh'], script_dir)
            command = 'cd ' + ut_dir + ';./get_child.sh'
            # print(command)
            run_script(command)
            download_file('get_alias.txt', local_dir)
            tcl_res_dict = {}
            with open('get_alias.txt', 'r') as fp:
                tcl_res_list = fp.readlines()
            for each_eqpt in tcl_res_list:
                each_alias_list = each_eqpt.split(' ')[0:-1]
                each_alias = each_alias_list[0][0:int(len(sys_name)) + 18]
                for each in each_alias_list:
                    if each_alias not in tcl_res_dict:
                        tcl_res_dict[each_alias] = set([each.replace(each_alias, '')])
                    else:
                        tcl_res_dict[each_alias].add(each.replace(each_alias, ''))
            # print(tcl_res_dict)
            ## 检查io check_dict中的点与env tcl_res_dict中的设备
            if check_dict.keys() ^ tcl_res_dict.keys() == set():
                print('---每个车站中的设备是相同的，继续测试！！！---')
                if error_code != 1:
                    error_code = 0
                ## 检查每个设备中的点是否一致
                for each_key in check_dict.keys():
                    if check_dict[each_key] - tcl_res_dict[each_key] == set():
                        print('---%s 设备中点的数据一致，测试通过！！！---'%(str(each_key)))
                        error_list.append(each_key + '设备中点的数据一致，测试通过！！！\n')
                        if error_code != 1:
                            error_code = 0
                    else:
                        if check_dict[each_key] - tcl_res_dict[each_key] != set():
                            print('---io中有，数据库无，请检查%s, %s设备中点的数据不一致，测试失败！！！---' % (str(check_dict[each_key] - tcl_res_dict[each_key]),str(each_key)))
                            error_list.append(str(each_key) + str(check_dict[each_key] - tcl_res_dict[each_key]) + 'io中有，数据库无，设备中点的数据不一致，测试失败！！！\n')
                            error_code = 1
                        if tcl_res_dict[each_key] - check_dict[each_key] != set():
                            print('---io中无，数据库有，请检查%s, %s设备中点的数据不一致，测试失败！！！---' % (
                            str(tcl_res_dict[each_key] - check_dict[each_key]), str(each_key)))
                            error_list.append(str(each_key) + str(
                                tcl_res_dict[each_key] - check_dict[each_key]) + 'io中无，数据库有，设备中点的数据不一致，测试失败！！！\n')
                            error_code = 1
            else:
                if check_dict.keys() - tcl_res_dict.keys() != set():
                    print('IO list中有: ' + str(check_dict.keys() - tcl_res_dict.keys()) + ',而生成DB中无, 请检查！！！')
                    error_list.append('IO list中有: ' + str(check_dict.keys() - tcl_res_dict.keys()) + ',而生成DB中无, 请检查！！！\n')
                    error_code = 1
                if tcl_res_dict.keys() - check_dict.keys() != set():
                    print('生成DB中有: ' + str(tcl_res_dict.keys() - check_dict.keys()) + ',而IO list中无, 请检查！！！')
                    error_list.append('生成DB中有: ' + str(tcl_res_dict.keys() - check_dict.keys()) + ',而IO list中无, 请检查！！！\n')
                    error_code = 1
            with open('eqp_pt_check.txt', 'w', encoding='utf-8') as f:
                f.writelines(error_list)
            print(txt_name)
            upload_file([txt_name], io_dir)

    # gen_txt(file_temp)
    # print('-----IO list.txt for ' + file_temp + ' has been created Successfully! -----')

    if error_code == 0:
        # Step 2. 上传文件到Server中
        upload_file(['save_new.sh', 'save_old.sh', 'add_right.sh', 'comstate.tcl', 'pt_sim.tcl',  'get_alias.txt', 'get_alias.tcl'], script_dir)

        print('-----.sh script files are uploaded to server successfully! -----')
        # Step 3.在Server中，执行.py文件去选取每一个点
        # command = 'cd /export/home/nbl2/OCC_A1TEMP2_181023/log/unit_test; ./abc.py'
        # run_script(command)
        # 首先set 通讯状态点

        window_to_unix()
        upload_file(['set_com.sh_template', 'set_pt.sh_template', 'run_test_new.py'], script_dir)
        command = 'cd ' + ut_dir + ';python run_test_new.py 2'
        run_script(command)
        download_file('out.txt', local_dir)
        download_file('test_result.txt', local_dir)
        print('执行完成')
        time.sleep(6000)
    else:
        exit('测试不通过，请查看eqp_pt_check.txt')
        time.sleep(6000)
